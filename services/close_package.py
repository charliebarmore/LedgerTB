"""Close package: everything a reviewer needs to tie out a finished period.

Two formats from the same underlying data — an Excel workbook for further
work, and a single multi-section PDF for the permanent file: Summary, Income
Statement, Balance Sheet, final Trial Balance (with the worksheet columns),
Transactions (every journal line in the period), Adjusting Entries, and
Receipts & Disbursements per cash account.
"""
from dataclasses import dataclass
from contextlib import contextmanager
from datetime import date, datetime
from html import unescape
from io import BytesIO
from typing import Dict, List, Optional

import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Border, Font, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.utils import ImageReader
from reportlab.platypus import (
    Image,
    KeepTogether,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from database import connection as dbconn
from constants import AccountSubtype
from database.connection import get_connection, get_cursor
from models.reports import (
    CASH_FLOW_STATEMENT_SECTIONS,
    ReportGenerator,
    TrialBalanceWorksheetRow,
)
from money import to_dollars
from services.branding import (
    ClientBranding,
    FirmBranding,
    get_branding,
    get_client_branding,
)
from utils.dates import long_date, long_datetime
from utils.fiscal_dates import fiscal_year_bounds
from utils.export import set_excel_literal

_HEADER_FONT = Font(bold=True)
_MONEY_FMT = "#,##0.00"
_STATEMENT_MONEY_FMT = '#,##0.00;(#,##0.00);"-"'
_STATEMENT_PERCENT_FMT = '0.0"%";(0.0"%");"-"'
_TOTAL_BORDER = Border(top=Side(style="thin", color="000000"))


@contextmanager
def consistent_export_window():
    """Prevent ledger writes while a multi-format close package is rendered.

    PDF and Excel perform several read queries. A reserved SQLite transaction
    lets other readers continue but prevents a writer from committing between
    those queries, so both files describe one committed ledger state.
    """
    conn = get_connection()
    try:
        from utils import books
        if dbconn.writes_permitted() and books.is_local_book(dbconn.DATABASE_PATH):
            conn.execute("BEGIN IMMEDIATE")
        else:
            conn.execute("BEGIN")
        yield
    finally:
        conn.rollback()
        conn.close()


@dataclass
class CashActivityRow:
    account_number: str
    account_name: str
    beginning: float
    receipts: float
    disbursements: float

    @property
    def ending(self) -> float:
        return round(self.beginning + self.receipts - self.disbursements, 2)


@dataclass
class ClosePackageSnapshot:
    transactions: List[dict]
    cash: List[CashActivityRow]
    income_statement: Dict
    balance_sheet: Dict
    comparative_income_statement: Dict
    comparative_balance_sheet: Dict
    cash_flow: Dict
    comparative_cash_flow: Dict
    comparative_trial_balance: Dict
    close_map: Optional[Dict]
    client_branding: ClientBranding
    branding: FirmBranding
    generated_at: datetime


def get_period_transactions(client_id: int, period_start: date, period_end: date) -> List[dict]:
    """Every journal line in the period, in entry order."""
    with get_cursor() as cursor:
        cursor.execute(
            """
            SELECT je.entry_date, je.id AS entry_id, je.entry_type,
                   je.description, je.source_reference,
                   a.account_number, a.name AS account_name,
                   jel.debit, jel.credit, jel.memo
            FROM journal_entries je
            JOIN journal_entry_lines jel ON jel.journal_entry_id = je.id
            JOIN accounts a ON a.id = jel.account_id
            WHERE je.client_id = ?
              AND je.entry_date BETWEEN ? AND ?
            ORDER BY je.entry_date, je.id, jel.id
            """,
            (client_id, period_start.isoformat(), period_end.isoformat()),
        )
        rows = cursor.fetchall()
    return [
        {
            "entry_date": row["entry_date"],
            "entry_id": row["entry_id"],
            "entry_type": row["entry_type"],
            # Descriptions can arrive through imported/assistant-authored source
            # text with HTML entities already encoded.  The PDF layer must escape
            # display text for ReportLab, so leaving ``&amp;`` here would escape it
            # a second time and print the entity literally.  Normalize the export
            # snapshot without rewriting the historical ledger value.
            "description": unescape(row["description"] or ""),
            "source_reference": row["source_reference"] or "",
            "account_number": row["account_number"],
            "account_name": row["account_name"],
            "debit": to_dollars(row["debit"]),
            "credit": to_dollars(row["credit"]),
            "memo": row["memo"] or "",
        }
        for row in rows
    ]


def get_cash_activity(client_id: int, period_start: date, period_end: date) -> List[CashActivityRow]:
    """Beginning balance, receipts, disbursements, ending — per cash account.

    Receipts and disbursements are the debits and credits hitting each cash
    account in the period: the cash-basis view of the books from the bank's
    side, which is what a receipts-and-disbursements engagement summarizes.
    """
    with get_cursor() as cursor:
        cursor.execute(
            """
            SELECT a.account_number, a.name AS account_name, a.type, a.subtype,
                   COALESCE(SUM(CASE WHEN je.entry_date < ?
                       OR (je.entry_type = 'Beginning Balance'
                           AND je.entry_date <= ?)
                       THEN jel.debit - jel.credit ELSE 0 END), 0) AS beginning,
                   COALESCE(SUM(CASE WHEN je.entry_date BETWEEN ? AND ?
                       AND je.entry_type != 'Beginning Balance'
                       THEN jel.debit ELSE 0 END), 0) AS receipts,
                   COALESCE(SUM(CASE WHEN je.entry_date BETWEEN ? AND ?
                       AND je.entry_type != 'Beginning Balance'
                       THEN jel.credit ELSE 0 END), 0) AS disbursements
            FROM accounts a
            LEFT JOIN journal_entry_lines jel ON jel.account_id = a.id
            LEFT JOIN journal_entries je ON je.id = jel.journal_entry_id
            WHERE a.client_id = ? AND a.type = 'Asset'
            GROUP BY a.id
            ORDER BY a.account_number
            """,
            (
                period_start.isoformat(),
                period_end.isoformat(),
                period_start.isoformat(), period_end.isoformat(),
                period_start.isoformat(), period_end.isoformat(),
                client_id,
            ),
        )
        rows = cursor.fetchall()
    return [
        CashActivityRow(
            account_number=row["account_number"],
            account_name=row["account_name"],
            beginning=to_dollars(row["beginning"]),
            receipts=to_dollars(row["receipts"]),
            disbursements=to_dollars(row["disbursements"]),
        )
        for row in rows
        if AccountSubtype.is_cash_like(
            row["type"], row["subtype"], row["account_name"]
        )
    ]


def load_close_package_snapshot(
    client_id: int, period_start: date, period_end: date
) -> ClosePackageSnapshot:
    """Capture the non-TB inputs once so PDF and Excel cannot drift apart."""
    close_map = None
    with get_cursor() as cursor:
        year = cursor.execute(
            "SELECT id FROM fiscal_periods WHERE client_id = ? AND period_type = 'Year' "
            "AND start_date = ? AND end_date = ? ORDER BY id DESC LIMIT 1",
            (client_id, period_start.isoformat(), period_end.isoformat()),
        ).fetchone()
    if year:
        from models.close_map import readiness
        close_map = readiness(client_id, year["id"])
    cash_flow = ReportGenerator.cash_flow_statement(
        client_id, period_start, period_end
    )
    comparative_cash_flow = ReportGenerator.comparative_cash_flow_statement(
        client_id, period_start, period_end, current_report=cash_flow
    )
    return ClosePackageSnapshot(
        transactions=get_period_transactions(client_id, period_start, period_end),
        cash=get_cash_activity(client_id, period_start, period_end),
        income_statement=ReportGenerator.income_statement(
            client_id, period_start, period_end
        ),
        balance_sheet=ReportGenerator.balance_sheet(client_id, period_end),
        comparative_income_statement=ReportGenerator.comparative_income_statement(
            client_id, period_start, period_end
        ),
        comparative_balance_sheet=ReportGenerator.comparative_balance_sheet(
            client_id, period_end
        ),
        cash_flow=cash_flow,
        comparative_cash_flow=comparative_cash_flow,
        comparative_trial_balance=ReportGenerator.comparative_trial_balance(
            client_id, period_end
        ),
        close_map=close_map,
        client_branding=get_client_branding(client_id),
        branding=get_branding(),
        generated_at=datetime.now(),
    )


def _earnings_tie_out(
    client_id: int, period_start: date, period_end: date,
    income_statement: Dict, balance_sheet: Dict,
) -> str:
    """Income-statement net income must equal the balance sheet's Current
    Year Earnings line for any fiscal-year-to-date package — the period
    starts at the fiscal year and ends inside it (docs/EARNINGS-ATTRIBUTION.md).
    A failing tie means the two statements disagree about the period's
    earnings — never export that silently."""
    with get_cursor() as cursor:
        row = cursor.execute(
            "SELECT fiscal_year_end_month FROM clients WHERE id = ?",
            (client_id,),
        ).fetchone()
    fye_month = (row["fiscal_year_end_month"]
                 if row and row["fiscal_year_end_month"] else 12)
    fy_start, fy_end = fiscal_year_bounds(period_end, fye_month)
    if period_start != fy_start or period_end > fy_end:
        return "Not a fiscal year-to-date period - not compared"
    current_year_earnings = sum(
        item["balance"] for item in balance_sheet["equity"]
        if item["name"] == "Current Year Earnings"
        and not item["account_number"]
    )
    if abs(income_statement["net_income"] - current_year_earnings) < 0.01:
        return "YES"
    return (
        "OUT OF BALANCE - net income "
        f"{income_statement['net_income']:,.2f} vs current year earnings "
        f"{current_year_earnings:,.2f}"
    )


def _append_literal_row(ws, values):
    """Append a row while ensuring every string is stored as literal text."""
    ws.append(values)
    cells = ws[ws.max_row]
    for cell, value in zip(cells, values):
        set_excel_literal(cell, value)
    return cells


def _write_table(ws, headers, data_rows, money_cols):
    money_cols = tuple(money_cols)
    _append_literal_row(ws, headers)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
    for row in data_rows:
        _append_literal_row(ws, row)
    for col_idx in money_cols:
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.number_format = _MONEY_FMT
    # Size every populated column from its actual role and contents. The old
    # fixed widths assumed column 2 was always Account Name, which made Entry #
    # enormous on transaction sheets while later money headers were clipped.
    for col_idx in range(1, ws.max_column + 1):
        values = (
            ws.cell(row=row_idx, column=col_idx).value
            for row_idx in range(1, ws.max_row + 1)
        )
        longest = max((len(str(value)) for value in values
                       if value is not None), default=0)
        minimum = 13 if col_idx in money_cols else 10
        width = min(max(longest + 2, minimum), 40)
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = width


def _statement_label(item: dict) -> str:
    """Account label shared by the two financial-statement sheets."""
    number = item.get("account_number") or ""
    return f"{number} - {item['name']}" if number else item["name"]


def _start_statement_sheet(wb, title: str, client_name: str, period_label: str,
                           accent_hex: str = ""):
    ws = wb.create_sheet(title)
    _append_literal_row(ws, [client_name, ""])
    ws["A1"].font = Font(
        bold=True, size=14, color=(accent_hex.lstrip("#") or "000000")
    )
    _append_literal_row(ws, [title, period_label])
    ws["A2"].font = Font(bold=True, size=12)
    ws["B2"].font = _HEADER_FONT
    _append_literal_row(ws, ["", ""])
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 18
    ws.sheet_view.showGridLines = False
    return ws


def _add_excel_logo(ws, logo: Optional[bytes], anchor: str,
                    max_width: float = 150, max_height: float = 55) -> None:
    """Add a safely scaled in-memory logo; bad image data cannot block export."""
    if not logo:
        return
    try:
        image = ExcelImage(BytesIO(logo))
        scale = min(max_width / image.width, max_height / image.height, 1)
        image.width *= scale
        image.height *= scale
        ws.add_image(image, anchor)
    except Exception:
        pass


def _append_statement_section(ws, title: str, items: List[dict],
                              total_label: str, total_value: float):
    section_cells = _append_literal_row(ws, [title.upper(), ""])
    for cell in section_cells:
        cell.font = _HEADER_FONT
    for item in items:
        cells = _append_literal_row(
            ws, [_statement_label(item), item["balance"]]
        )
        cells[1].number_format = _STATEMENT_MONEY_FMT
    total_cells = _append_literal_row(ws, [total_label, total_value])
    for cell in total_cells:
        cell.font = _HEADER_FONT
        cell.border = _TOTAL_BORDER
    total_cells[1].number_format = _STATEMENT_MONEY_FMT


def _append_statement_total(ws, label: str, value: float):
    cells = _append_literal_row(ws, [label, value])
    for cell in cells:
        cell.font = _HEADER_FONT
        cell.border = _TOTAL_BORDER
    cells[1].number_format = _STATEMENT_MONEY_FMT


def _comparison_values(item: Dict) -> list:
    return [
        item['current'], item['prior'], item['change'], item['change_percent']
    ]


def _prepare_comparative_statement_sheet(ws, current_label: str, prior_label: str):
    cells = _append_literal_row(
        ws, ["", current_label, prior_label, "$ Change", "% Change"]
    )
    for cell in cells:
        cell.font = _HEADER_FONT
    for column in ("B", "C", "D"):
        ws.column_dimensions[column].width = 18
    ws.column_dimensions["E"].width = 13


def _append_comparative_statement_section(
    ws, title: str, items: List[dict], total_label: str, total: Dict
):
    section_cells = _append_literal_row(ws, [title.upper(), "", "", "", ""])
    for cell in section_cells:
        cell.font = _HEADER_FONT
    for item in items:
        cells = _append_literal_row(
            ws, [_statement_label(item)] + _comparison_values(item)
        )
        for cell in cells[1:4]:
            cell.number_format = _STATEMENT_MONEY_FMT
        cells[4].number_format = _STATEMENT_PERCENT_FMT
    _append_comparative_statement_total(ws, total_label, total)


def _append_comparative_statement_groups(
    ws, title: str, groups: List[Dict], total_label: str, total: Dict
):
    """Render statement groups without changing the established total rows."""
    section_cells = _append_literal_row(ws, [title.upper(), "", "", "", ""])
    for cell in section_cells:
        cell.font = _HEADER_FONT

    if not groups:
        _append_literal_row(ws, [f"No {title.lower()} recorded", "", "", "", ""])

    for group in groups:
        group_cells = _append_literal_row(
            ws, [f"  {group['group']}", "", "", "", ""]
        )
        for cell in group_cells:
            cell.font = _HEADER_FONT
        for item in group['accounts']:
            cells = _append_literal_row(
                ws, [f"    {_statement_label(item)}"] + _comparison_values(item)
            )
            for cell in cells[1:4]:
                cell.number_format = _STATEMENT_MONEY_FMT
            cells[4].number_format = _STATEMENT_PERCENT_FMT
        subtotal_cells = _append_literal_row(
            ws,
            [f"  Total {group['group']}"] + _comparison_values(group['subtotal']),
        )
        for cell in subtotal_cells:
            cell.font = _HEADER_FONT
            cell.border = _TOTAL_BORDER
        for cell in subtotal_cells[1:4]:
            cell.number_format = _STATEMENT_MONEY_FMT
        subtotal_cells[4].number_format = _STATEMENT_PERCENT_FMT

    _append_comparative_statement_total(ws, total_label, total)


def _income_statement_rows(report: Dict) -> List[tuple]:
    """Compatibility wrapper around the shared statement row assembly."""
    return ReportGenerator.income_statement_rows(report)


def _append_comparative_income_statement(ws, report: Dict) -> None:
    for kind, label, values in _income_statement_rows(report):
        prefix = (
            '    ' if kind == 'item'
            else '  ' if kind in {'group', 'group_total'}
            else ''
        )
        cells = _append_literal_row(
            ws,
            [f"{prefix}{label}"] + (
                _comparison_values(values) if values is not None else ["", "", "", ""]
            ),
        )
        if kind in {'section', 'group', 'group_total', 'subtotal', 'total'}:
            for cell in cells:
                cell.font = _HEADER_FONT
        if kind in {'group_total', 'subtotal', 'total'}:
            for cell in cells:
                cell.border = _TOTAL_BORDER
        if values is not None:
            for cell in cells[1:4]:
                cell.number_format = _STATEMENT_MONEY_FMT
            cells[4].number_format = _STATEMENT_PERCENT_FMT


def _append_comparative_statement_total(ws, label: str, total: Dict):
    cells = _append_literal_row(ws, [label] + _comparison_values(total))
    for cell in cells:
        cell.font = _HEADER_FONT
        cell.border = _TOTAL_BORDER
    for cell in cells[1:4]:
        cell.number_format = _STATEMENT_MONEY_FMT
    cells[4].number_format = _STATEMENT_PERCENT_FMT


def build_close_package(
    client_id: int,
    client_name: str,
    period_start: date,
    period_end: date,
    tb_rows: List[TrialBalanceWorksheetRow],
    snapshot: Optional[ClosePackageSnapshot] = None,
) -> BytesIO:
    """Build the close-package workbook and return it as an in-memory file."""
    snapshot = snapshot or load_close_package_snapshot(
        client_id, period_start, period_end
    )
    transactions = snapshot.transactions
    ajes = [t for t in transactions if t["entry_type"] == "Adjusting"]
    cash = snapshot.cash
    income_statement = snapshot.income_statement
    balance_sheet = snapshot.balance_sheet
    comparative_income = snapshot.comparative_income_statement
    comparative_balance = snapshot.comparative_balance_sheet
    cash_flow = snapshot.cash_flow
    comparative_cash_flow = snapshot.comparative_cash_flow
    comparative_tb = snapshot.comparative_trial_balance
    close_map = snapshot.close_map
    client_branding = snapshot.client_branding
    firm_branding = snapshot.branding
    display_name = client_branding.display_name or client_name
    accent_hex = client_branding.accent_hex or firm_branding.accent_hex

    wb = openpyxl.Workbook()

    # ---- Summary
    ws = wb.active
    ws.title = "Summary"
    total_dr = round(sum(r.adjusted_dr for r in tb_rows), 2)
    total_cr = round(sum(r.adjusted_cr for r in tb_rows), 2)
    lines = [
        (display_name, ""),
        ("Close package", f"{period_start.isoformat()} to {period_end.isoformat()}"),
        ("Generated", snapshot.generated_at.strftime("%Y-%m-%d %H:%M")),
    ]
    if client_branding.tagline:
        lines.append(("Client", client_branding.tagline))
    if firm_branding.firm_name:
        lines.append(("Prepared by", firm_branding.firm_name))
    if firm_branding.tagline:
        lines.append(("Preparer details", firm_branding.tagline))
    lines += [
        ("", ""),
        ("Final trial balance - total debits", total_dr),
        ("Final trial balance - total credits", total_cr),
        ("In balance", "YES" if abs(total_dr - total_cr) < 0.01 else "OUT OF BALANCE"),
        ("Net income for period", income_statement["net_income"]),
        ("Prior-year net income",
         comparative_income["net_income"]["prior"]
         if comparative_income["prior_available"] else "No prior-year data"),
        ("Balance sheet - total assets", balance_sheet["total_assets"]),
        ("Prior-year total assets",
         comparative_balance["total_assets"]["prior"]
         if comparative_balance["prior_available"] else "No prior-year data"),
        ("Balance sheet - liabilities & equity",
         balance_sheet["total_liabilities_equity"]),
        ("Balance sheet in balance",
         "YES" if abs(balance_sheet["total_assets"] -
                      balance_sheet["total_liabilities_equity"]) < 0.01
         else "OUT OF BALANCE"),
        ("Net income ties to balance sheet earnings",
         _earnings_tie_out(client_id, period_start, period_end,
                           income_statement, balance_sheet)),
        ("Cash flow status",
         "READY" if cash_flow["ready"] else "REVIEW WARNINGS"),
        ("Cash flow - net change in cash", cash_flow["computed_cash_change"]),
        ("Journal lines in period", len(transactions)),
        ("Adjusting entry lines", len(ajes)),
        ("Close Map",
         (f"{close_map['reviewed_count']} of {close_map['required_count']} required balances reviewed"
          if close_map else "Not available for this period")),
        ("", ""),
        ("Cash accounts", "Receipts / Disbursements"),
    ]
    for row in cash:
        lines.append((
            f"{row.account_number} - {row.account_name}",
            f"{row.receipts:,.2f} / {row.disbursements:,.2f}",
        ))
    for left, right in lines:
        _append_literal_row(ws, [left, right])
    ws["A1"].font = Font(
        bold=True, size=14, color=(accent_hex.lstrip("#") or "000000")
    )
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 28
    ws.row_dimensions[1].height = 46
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["F"].width = 22
    _add_excel_logo(ws, client_branding.logo, "D1")
    _add_excel_logo(ws, firm_branding.logo, "F1", max_width=120, max_height=45)

    # ---- Income Statement
    ws = _start_statement_sheet(
        wb, "Income Statement", display_name,
        f"{period_start.isoformat()} to {period_end.isoformat()}",
        accent_hex,
    )
    _prepare_comparative_statement_sheet(
        ws,
        f"{period_start.isoformat()} to {period_end.isoformat()}",
        (f"{comparative_income['prior_period']['start'].isoformat()} to "
         f"{comparative_income['prior_period']['end'].isoformat()}"),
    )
    if not comparative_income['prior_available']:
        _append_literal_row(ws, ["No prior-year data", "", "", "", ""])
    _append_comparative_income_statement(ws, comparative_income)

    # ---- Balance Sheet
    ws = _start_statement_sheet(
        wb, "Balance Sheet", display_name, f"As of {period_end.isoformat()}",
        accent_hex,
    )
    _prepare_comparative_statement_sheet(
        ws,
        f"As of {period_end.isoformat()}",
        f"As of {comparative_balance['prior_as_of'].isoformat()}",
    )
    if not comparative_balance['prior_available']:
        _append_literal_row(ws, ["No prior-year data", "", "", "", ""])
    _append_comparative_statement_groups(
        ws, "Assets", comparative_balance["asset_groups"],
        "Total Assets", comparative_balance["total_assets"],
    )
    _append_literal_row(ws, ["", "", "", "", ""])
    _append_comparative_statement_groups(
        ws, "Liabilities", comparative_balance["liability_groups"],
        "Total Liabilities", comparative_balance["total_liabilities"],
    )
    _append_literal_row(ws, ["", "", "", "", ""])
    _append_comparative_statement_groups(
        ws, "Equity", comparative_balance["equity_groups"],
        "Total Equity", comparative_balance["total_equity"],
    )
    _append_literal_row(ws, ["", "", "", "", ""])
    _append_comparative_statement_total(
        ws, "TOTAL LIABILITIES & EQUITY",
        comparative_balance["total_liabilities_equity"],
    )
    _append_statement_total(
        ws, "BALANCE CHECK",
        round(balance_sheet["total_assets"] -
              balance_sheet["total_liabilities_equity"], 2),
    )

    # ---- Trial Balance (final, with the worksheet's supporting columns)
    ws = wb.create_sheet("Trial Balance")
    comparison_by_number = {
        row['account_number']: row for row in comparative_tb['accounts']
    }
    tb_export_rows = []
    current_numbers = set()
    for row in tb_rows:
        current_numbers.add(row.account_number)
        comparison_row = comparison_by_number.get(row.account_number, {})
        tb_export_rows.append([
            row.account_number, row.account_name, row.account_type,
            row.beginning_dr or None, row.beginning_cr or None,
            row.period_debits or None, row.period_credits or None,
            row.aje_debits or None, row.aje_credits or None,
            row.adjusted_dr or None, row.adjusted_cr or None,
            comparison_row.get('prior_debit') or None,
            comparison_row.get('prior_credit') or None,
        ])
    for comparison_row in comparative_tb['accounts']:
        if comparison_row['account_number'] in current_numbers:
            continue
        if not (comparison_row.get('prior_debit') or
                comparison_row.get('prior_credit')):
            continue
        tb_export_rows.append([
            comparison_row['account_number'], comparison_row['name'],
            comparison_row['type'], None, None, None, None, None, None,
            None, None, comparison_row.get('prior_debit') or None,
            comparison_row.get('prior_credit') or None,
        ])
    _write_table(
        ws,
        ["Acct #", "Account Name", "Type", "Beg Dr", "Beg Cr",
         "Activity Dr", "Activity Cr", "AJE Dr", "AJE Cr",
         "Final Dr", "Final Cr", "PY Final Dr", "PY Final Cr"],
        tb_export_rows,
        money_cols=range(4, 14),
    )
    totals = ws.max_row + 1
    ws.cell(row=totals, column=1, value="TOTALS").font = _HEADER_FONT
    # Computed values, not =SUM() formulas: openpyxl writes no cached results,
    # so a formula cell reads as literal "=SUM(...)" text to anything that
    # opens the workbook without a calc engine (LedgerPDF renders exactly what
    # the file says) until a human opens and re-saves it in Excel.
    _tb_totals = [
        round(sum(r.beginning_dr for r in tb_rows), 2),
        round(sum(r.beginning_cr for r in tb_rows), 2),
        round(sum(r.period_debits for r in tb_rows), 2),
        round(sum(r.period_credits for r in tb_rows), 2),
        round(sum(r.aje_debits for r in tb_rows), 2),
        round(sum(r.aje_credits for r in tb_rows), 2),
        round(sum(r.adjusted_dr for r in tb_rows), 2),
        round(sum(r.adjusted_cr for r in tb_rows), 2),
        comparative_tb['prior_total_debits'],
        comparative_tb['prior_total_credits'],
    ]
    for col_idx, value in zip(range(4, 14), _tb_totals):
        cell = ws.cell(row=totals, column=col_idx, value=value)
        cell.font = _HEADER_FONT
        cell.number_format = _MONEY_FMT

    # ---- Close Map (annual packages only)
    if close_map is not None:
        ws = wb.create_sheet("Close Map")
        _write_table(
            ws,
            ["Acct #", "Account", "Type", "Adjusted", "Prior Year", "$ Change",
             "% Change", "Lead Sheet", "Evidence", "Evidence References",
             "Open Notes", "Open Review Notes", "Status",
             "Prepared By", "Prepared At", "Reviewed By", "Reviewed At",
             "Explanation", "Exclusion Reason"],
            [
                [row.account_number, row.account_name, row.account_type,
                 row.current_balance, row.prior_balance, row.change,
                 row.change_percent,
                 (f"{row.group_code} - {row.group_name}" if row.group_code else ""),
                 row.evidence_count, row.evidence_references,
                 row.open_note_count, row.open_notes, row.status,
                 row.prepared_by, row.prepared_at, row.reviewed_by, row.reviewed_at,
                 row.explanation, row.exclusion_reason]
                for row in close_map["rows"]
            ],
            money_cols=(4, 5, 6),
        )
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["J"].width = 45
        ws.column_dimensions["L"].width = 60
        ws.column_dimensions["R"].width = 60
        ws.column_dimensions["S"].width = 45

    # ---- Transactions
    ws = wb.create_sheet("Transactions")
    _write_table(
        ws,
        ["Date", "Entry #", "Type", "Description", "Acct #", "Account",
         "Debit", "Credit", "Memo", "Source Ref"],
        [
            [t["entry_date"], t["entry_id"], t["entry_type"], t["description"],
             t["account_number"], t["account_name"],
             t["debit"] or None, t["credit"] or None, t["memo"],
             t["source_reference"]]
            for t in transactions
        ],
        money_cols=(7, 8),
    )
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["F"].width = 30

    # ---- Adjusting Entries
    ws = wb.create_sheet("Adjusting Entries")
    _write_table(
        ws,
        ["Date", "Entry #", "AJE Ref", "Description", "Acct #", "Account",
         "Debit", "Credit", "Memo"],
        [
            [t["entry_date"], t["entry_id"], t["source_reference"],
             t["description"], t["account_number"], t["account_name"],
             t["debit"] or None, t["credit"] or None, t["memo"]]
            for t in ajes
        ],
        money_cols=(7, 8),
    )
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["F"].width = 30

    # ---- Receipts & Disbursements
    ws = wb.create_sheet("Receipts & Disbursements")
    _write_table(
        ws,
        ["Acct #", "Cash Account", "Beginning", "Total Receipts",
         "Total Disbursements", "Ending"],
        [
            [row.account_number, row.account_name, row.beginning,
             row.receipts, row.disbursements, row.ending]
            for row in cash
        ],
        money_cols=range(3, 7),
    )
    if cash:
        totals = ws.max_row + 1
        ws.cell(row=totals, column=2, value="TOTALS").font = _HEADER_FONT
        _cash_totals = [
            round(sum(r.beginning for r in cash), 2),
            round(sum(r.receipts for r in cash), 2),
            round(sum(r.disbursements for r in cash), 2),
            round(sum(r.ending for r in cash), 2),
        ]
        for col_idx, value in zip(range(3, 7), _cash_totals):
            cell = ws.cell(row=totals, column=col_idx, value=value)
            cell.font = _HEADER_FONT
            cell.number_format = _MONEY_FMT

    # ---- Cash Flow (appended so all established sheet positions stay stable)
    ws = _start_statement_sheet(
        wb, "Cash Flow", display_name,
        f"{period_start.isoformat()} to {period_end.isoformat()}",
        accent_hex,
    )
    _prepare_comparative_statement_sheet(
        ws,
        f"{period_start.isoformat()} to {period_end.isoformat()}",
        (f"{comparative_cash_flow['prior_period']['start'].isoformat()} to "
         f"{comparative_cash_flow['prior_period']['end'].isoformat()}"),
    )
    if not comparative_cash_flow['prior_available']:
        _append_literal_row(ws, ["No prior-year data", "", "", "", ""])
    cash_flow_sections = CASH_FLOW_STATEMENT_SECTIONS + ((
        "Unclassified Cash Activity",
        "unclassified",
        "Net Unclassified Cash Activity",
    ),)
    for title, key, total_label in cash_flow_sections:
        section = comparative_cash_flow[key]
        if (
            key == "unclassified"
            and not section["lines"]
            and not section["current_entries"]
            and not section["prior_entries"]
        ):
            continue
        _append_comparative_statement_section(
            ws, title, section["lines"], total_label, section["total"]
        )
        _append_literal_row(ws, ["", "", "", "", ""])
    _append_comparative_statement_total(
        ws, "NET CHANGE IN CASH", comparative_cash_flow["computed_cash_change"]
    )
    reconciliation = comparative_cash_flow["reconciliation_difference"]
    if reconciliation["current"] or reconciliation["prior"]:
        _append_comparative_statement_total(
            ws, "CASH FLOW RECONCILIATION DIFFERENCE", reconciliation
        )
    _append_comparative_statement_total(
        ws, "CASH AT BEGINNING OF PERIOD", comparative_cash_flow["cash_beginning"]
    )
    _append_comparative_statement_total(
        ws, "CASH AT END OF PERIOD", comparative_cash_flow["cash_ending"]
    )
    _append_literal_row(ws, [
        "STATUS",
        ("READY" if comparative_cash_flow["current_ready"] else "REVIEW WARNINGS"),
        (
            "" if not comparative_cash_flow["prior_available"] else
            ("READY" if comparative_cash_flow["prior_ready"] else "REVIEW WARNINGS")
        ),
        "", "",
    ])
    for warning in comparative_cash_flow["current_warnings"]:
        _append_literal_row(ws, [f"Current warning: {warning}", "", "", "", ""])
    for warning in comparative_cash_flow["prior_warnings"]:
        _append_literal_row(ws, [f"Prior-year warning: {warning}", "", "", "", ""])
    for period_name, entries, amount_column in [
        ("CURRENT UNCLASSIFIED ENTRY DETAILS",
         comparative_cash_flow["unclassified"]["current_entries"], 2),
        ("PRIOR-YEAR UNCLASSIFIED ENTRY DETAILS",
         comparative_cash_flow["unclassified"]["prior_entries"], 3),
    ]:
        if not entries:
            continue
        _append_literal_row(ws, ["", "", "", "", ""])
        heading_cells = _append_literal_row(ws, [period_name, "", "", "", ""])
        for cell in heading_cells:
            cell.font = _HEADER_FONT
        for entry in entries:
            accounts = ", ".join(entry["account_numbers"]) or "none"
            cells = _append_literal_row(ws, [
                (f"{entry['entry_date']} | Entry #{entry['entry_id']} | "
                 f"{entry['reason']} | {entry['description'] or 'No description'} | "
                 f"Accounts {accounts}"),
                "", "", "", "",
            ])
            cells[amount_column - 1].value = entry["amount"]
            cells[amount_column - 1].number_format = _STATEMENT_MONEY_FMT

    for period_name, entries, amount_column in [
        ("CURRENT NONCASH INVESTING AND FINANCING ACTIVITY",
         comparative_cash_flow["current_noncash_items"], 2),
        ("PRIOR-YEAR NONCASH INVESTING AND FINANCING ACTIVITY",
         comparative_cash_flow["prior_noncash_items"], 3),
    ]:
        if not entries:
            continue
        _append_literal_row(ws, ["", "", "", "", ""])
        heading_cells = _append_literal_row(ws, [period_name, "", "", "", ""])
        for cell in heading_cells:
            cell.font = _HEADER_FONT
        for entry in entries:
            accounts = ", ".join(entry["accounts"]) or "none"
            cells = _append_literal_row(ws, [
                (f"{entry['entry_date']} | Entry #{entry['entry_id']} | "
                 f"{entry['description'] or 'No description'} | "
                 f"Accounts {accounts}"),
                "", "", "", "",
            ])
            cells[amount_column - 1].value = entry["amount"]
            cells[amount_column - 1].number_format = _STATEMENT_MONEY_FMT

    for sheet in wb.worksheets:
        sheet.oddFooter.left.text = display_name
        sheet.oddFooter.center.text = (
            f"Prepared by {firm_branding.firm_name}"
            if firm_branding.firm_name else "Prepared by LedgerTB"
        )
        sheet.oddFooter.right.text = "Page &P of &N"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# --------------------------------------------------------------------- PDF

_PDF_BODY = ParagraphStyle("body", fontName="Helvetica", fontSize=8, leading=10)
_PDF_H1 = ParagraphStyle("h1", fontName="Helvetica-Bold", fontSize=16, leading=20,
                         spaceAfter=6)
_PDF_H2 = ParagraphStyle("h2", fontName="Helvetica-Bold", fontSize=12, leading=15,
                         spaceBefore=6, spaceAfter=8)
_PDF_META = ParagraphStyle("meta", fontName="Helvetica", fontSize=10, leading=14,
                           textColor=colors.HexColor("#444444"))


def _money(value: float) -> str:
    if not value:
        return ""
    body = f"{abs(value):,.2f}"
    return f"({body})" if value < 0 else body


def _money_total(value: float) -> str:
    """Display statement totals explicitly, including a meaningful zero."""
    body = f"{abs(value):,.2f}"
    return f"({body})" if value < 0 else body


def _percent(value: Optional[float]) -> str:
    if value is None:
        return ""
    body = f"{abs(value):,.1f}%"
    return f"({body})" if value < 0 else body


def _pdf_comparison_values(item: Dict, totals: bool = False) -> list:
    money = _money_total if totals else _money
    return [
        money(item['current']),
        "" if item['prior'] is None else money(item['prior']),
        "" if item['change'] is None else money(item['change']),
        _percent(item['change_percent']),
    ]


def _wrap(text: str) -> Paragraph:
    return Paragraph((text or "").replace("&", "&amp;").replace("<", "&lt;"), _PDF_BODY)


def _safe_paragraph(text: str, style: ParagraphStyle) -> Paragraph:
    escaped = (text or "").replace("&", "&amp;").replace("<", "&lt;")
    return Paragraph(escaped, style)


def _pdf_table(headers, data_rows, col_widths, money_from: Optional[int],
               totals_row=None, bold_data_rows=None,
               ruled_data_rows=None, no_split_data_ranges=None) -> Table:
    """A report table: bold repeating header, right-aligned money columns."""
    rows = [headers] + data_rows
    if totals_row is not None:
        rows.append(totals_row)
    table = Table(rows, colWidths=col_widths, repeatRows=1, hAlign="LEFT")
    style = [
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("LINEBELOW", (0, 0), (-1, 0), 0.75, colors.black),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#F4F4F0")]),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]
    if money_from is not None:
        style.append(("ALIGN", (money_from, 0), (-1, -1), "RIGHT"))
    if totals_row is not None:
        style += [
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("LINEABOVE", (0, -1), (-1, -1), 0.75, colors.black),
        ]
    for data_index in bold_data_rows or ():
        table_row = data_index + 1
        style.append((
            "FONTNAME", (0, table_row), (-1, table_row), "Helvetica-Bold"
        ))
    for data_index in ruled_data_rows or ():
        table_row = data_index + 1
        style.append((
            "LINEABOVE", (0, table_row), (-1, table_row),
            0.5, colors.HexColor("#777777"),
        ))
    for start_index, end_index in no_split_data_ranges or ():
        style.append((
            "NOSPLIT", (0, start_index + 1), (-1, end_index + 1)
        ))
    table.setStyle(TableStyle(style))
    return table


def _pdf_income_statement_table(report: Dict) -> Table:
    data_rows = []
    bold_rows = []
    ruled_rows = []
    for kind, label, values in _income_statement_rows(report):
        prefix = (
            '    ' if kind == 'item'
            else '  ' if kind in {'group', 'group_total'}
            else ''
        )
        row = [
            _wrap(label) if kind == 'item' else f"{prefix}{label}"
        ]
        row += (
            _pdf_comparison_values(
                values, totals=kind in {'group_total', 'subtotal', 'total'}
            )
            if values is not None else ["", "", "", ""]
        )
        data_rows.append(row)
        row_index = len(data_rows) - 1
        if kind in {'section', 'group', 'group_total', 'subtotal', 'total'}:
            bold_rows.append(row_index)
        if kind in {'group_total', 'subtotal', 'total'}:
            ruled_rows.append(row_index)
    return _pdf_table(
        ["Account", "Current", "Prior Year", "$ Change", "% Change"],
        data_rows,
        [4.4 * inch, 1.35 * inch, 1.35 * inch, 1.35 * inch, 1.0 * inch],
        money_from=1,
        bold_data_rows=bold_rows,
        ruled_data_rows=ruled_rows,
    )


def _pdf_grouped_comparison_table(groups: List[Dict], empty_label: str,
                                  total_label: str, total: Dict) -> Table:
    data_rows = []
    bold_rows = []
    ruled_rows = []
    no_split_ranges = []
    if not groups:
        data_rows.append([empty_label, "", "", "", ""])
    for group in groups:
        group_start = len(data_rows)
        data_rows.append([f"  {group['group']}", "", "", "", ""])
        bold_rows.append(len(data_rows) - 1)
        for item in group['accounts']:
            data_rows.append([
                _wrap(_statement_label(item)),
                *_pdf_comparison_values(item),
            ])
        if group['accounts']:
            no_split_ranges.append((group_start, group_start + 1))
        data_rows.append([
            f"  Total {group['group']}",
            *_pdf_comparison_values(group['subtotal'], totals=True),
        ])
        bold_rows.append(len(data_rows) - 1)
        ruled_rows.append(len(data_rows) - 1)
    return _pdf_table(
        ["Account", "Current", "Prior Year", "$ Change", "% Change"],
        data_rows,
        [4.4 * inch, 1.35 * inch, 1.35 * inch, 1.35 * inch, 1.0 * inch],
        money_from=1,
        totals_row=[total_label] + _pdf_comparison_values(total, totals=True),
        bold_data_rows=bold_rows,
        ruled_data_rows=ruled_rows,
        no_split_data_ranges=no_split_ranges,
    )


def _logo_flowable(identity, max_height: float):
    """A stored brand logo scaled to the masthead, or None."""
    if not identity.logo:
        return None
    try:
        reader = ImageReader(BytesIO(identity.logo))
        width, height = reader.getSize()
        scale = max_height / float(height)
        return Image(BytesIO(identity.logo),
                     width=width * scale, height=max_height)
    except Exception:
        return None  # a corrupt logo must never block the close package


def build_close_package_pdf(
    client_id: int,
    client_name: str,
    period_start: date,
    period_end: date,
    tb_rows: List[TrialBalanceWorksheetRow],
    snapshot: Optional[ClosePackageSnapshot] = None,
) -> BytesIO:
    """One presentable PDF: Summary, statements, TB, transactions, AJEs."""
    snapshot = snapshot or load_close_package_snapshot(
        client_id, period_start, period_end
    )
    transactions = snapshot.transactions
    ajes = [t for t in transactions if t["entry_type"] == "Adjusting"]
    cash = snapshot.cash
    income_statement = snapshot.income_statement
    balance_sheet = snapshot.balance_sheet
    comparative_income = snapshot.comparative_income_statement
    comparative_balance = snapshot.comparative_balance_sheet
    cash_flow = snapshot.cash_flow
    comparative_cash_flow = snapshot.comparative_cash_flow
    comparative_tb = snapshot.comparative_trial_balance
    close_map = snapshot.close_map
    period_label = f"{long_date(period_start)} to {long_date(period_end)}"

    client_branding = snapshot.client_branding
    firm_branding = snapshot.branding
    display_name = client_branding.display_name or client_name
    accent_hex = client_branding.accent_hex or firm_branding.accent_hex
    accent = colors.HexColor(accent_hex) if accent_hex else colors.black
    heading_1 = ParagraphStyle("bh1", parent=_PDF_H1, textColor=accent)
    heading_2 = ParagraphStyle(
        "bh2", parent=_PDF_H2, textColor=accent, keepWithNext=True
    )

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(letter),
        leftMargin=0.5 * inch, rightMargin=0.5 * inch,
        topMargin=0.6 * inch, bottomMargin=0.55 * inch,
        title=f"Close Package - {display_name}",
        author=firm_branding.firm_name or "LedgerTB",
    )

    footer_left = f"{display_name} - {period_label}"
    if firm_branding.firm_name:
        footer_left = f"{footer_left} · Prepared by {firm_branding.firm_name}"

    def _footer(canvas, _doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7.5)
        canvas.setFillColor(colors.HexColor("#666666"))
        canvas.drawString(0.5 * inch, 0.3 * inch, footer_left)
        canvas.drawRightString(
            doc.pagesize[0] - 0.5 * inch, 0.3 * inch, f"Page {canvas.getPageNumber()}"
        )
        canvas.restoreState()

    total_dr = round(sum(r.adjusted_dr for r in tb_rows), 2)
    total_cr = round(sum(r.adjusted_cr for r in tb_rows), 2)
    balanced = abs(total_dr - total_cr) < 0.01

    masthead = []
    client_logo = _logo_flowable(client_branding, max_height=0.6 * inch)
    if client_logo:
        client_logo.hAlign = "LEFT"
        masthead += [client_logo, Spacer(1, 6)]

    story = masthead + [_safe_paragraph(display_name, heading_1)]
    if client_branding.tagline:
        story += [_safe_paragraph(client_branding.tagline, _PDF_META)]
    story += [
        Paragraph("Close Package", _PDF_META),
        Paragraph(period_label, _PDF_META),
        Paragraph(f"Generated {long_datetime(snapshot.generated_at)}", _PDF_META),
        Spacer(1, 10),
    ]
    firm_logo = _logo_flowable(firm_branding, max_height=0.32 * inch)
    if firm_logo:
        firm_logo.hAlign = "LEFT"
        story += [firm_logo, Spacer(1, 4)]
    if firm_branding.firm_name:
        firm_line = "Prepared by " + firm_branding.firm_name
        if firm_branding.tagline:
            firm_line += f" · {firm_branding.tagline}"
        story += [_safe_paragraph(firm_line, _PDF_META)]
    story += [
        Spacer(1, 18),
        Paragraph("Summary", heading_2),
        _pdf_table(
            ["", ""],
            [
                ["Final trial balance - total debits", _money(total_dr)],
                ["Final trial balance - total credits", _money(total_cr)],
                ["In balance", "Yes" if balanced else "OUT OF BALANCE"],
                ["Net income ties to balance sheet earnings",
                 _earnings_tie_out(client_id, period_start, period_end,
                                   income_statement, balance_sheet)],
                ["Journal lines in period", str(len(transactions))],
                ["Adjusting entry lines", str(len(ajes))],
                ["Close Map",
                 (f"{close_map['reviewed_count']} of {close_map['required_count']} required balances reviewed"
                  if close_map else "Not available for this period")],
            ],
            [3.4 * inch, 1.6 * inch], money_from=1,
        ),
        Spacer(1, 14),
        Paragraph("Cash Activity", heading_2),
        _pdf_table(
            ["Acct #", "Cash Account", "Beginning", "Total Receipts",
             "Total Disbursements", "Ending"],
            [[r.account_number, _wrap(r.account_name), _money(r.beginning),
              f"{r.receipts:,.2f}", f"{r.disbursements:,.2f}", _money(r.ending)]
             for r in cash] or [["", "No cash accounts", "", "", "", ""]],
            [0.6 * inch, 2.6 * inch, 1.1 * inch, 1.2 * inch, 1.5 * inch, 1.1 * inch],
            money_from=2,
            totals_row=(
                ["", "TOTALS",
                 _money(round(sum(r.beginning for r in cash), 2)),
                 f"{sum(r.receipts for r in cash):,.2f}",
                 f"{sum(r.disbursements for r in cash):,.2f}",
                 _money(round(sum(r.ending for r in cash), 2))]
                if cash else None
            ),
        ),
        PageBreak(),
    ]

    # ---- Income Statement
    story += [
        Paragraph("Income Statement", heading_2),
        Paragraph(period_label, _PDF_META),
        Paragraph(
            (f"Prior period: {long_date(comparative_income['prior_period']['start'])} "
             f"to {long_date(comparative_income['prior_period']['end'])}")
            if comparative_income['prior_available'] else "No prior-year data",
            _PDF_META,
        ),
        Spacer(1, 12),
        _pdf_income_statement_table(comparative_income),
        PageBreak(),
    ]

    # ---- Balance Sheet
    story += [
        Paragraph("Balance Sheet", heading_2),
        Paragraph(f"As of {long_date(period_end)}", _PDF_META),
        Paragraph(
            f"Prior year as of {long_date(comparative_balance['prior_as_of'])}"
            if comparative_balance['prior_available'] else "No prior-year data",
            _PDF_META,
        ),
        Spacer(1, 12),
    ]
    for section_title, items, total_label, total_value in [
        ("Assets", comparative_balance["asset_groups"],
         "Total Assets", comparative_balance["total_assets"]),
        ("Liabilities", comparative_balance["liability_groups"],
         "Total Liabilities", comparative_balance["total_liabilities"]),
        ("Equity", comparative_balance["equity_groups"],
         "Total Equity", comparative_balance["total_equity"]),
    ]:
        section_block = [
            Paragraph(section_title, heading_2),
            _pdf_grouped_comparison_table(
                items,
                f"No {section_title.lower()} recorded",
                total_label,
                total_value,
            ),
            Spacer(1, 10),
        ]
        section_row_count = 1 + sum(
            2 + len(group['accounts']) for group in items
        )
        if section_row_count <= 12:
            story.append(KeepTogether(section_block))
        else:
            story.extend(section_block)
    balance_difference = round(
        balance_sheet["total_assets"] -
        balance_sheet["total_liabilities_equity"], 2
    )
    story += [
        _pdf_table(
            ["", "Current", "Prior Year", "$ Change", "% Change"], [],
            [4.4 * inch, 1.35 * inch, 1.35 * inch, 1.35 * inch, 1.0 * inch],
            money_from=1,
            totals_row=["TOTAL LIABILITIES & EQUITY"] + _pdf_comparison_values(
                comparative_balance["total_liabilities_equity"], totals=True
            ),
        ),
        Spacer(1, 8),
        Paragraph(
            "Balance sheet is in balance."
            if abs(balance_difference) < 0.01
            else f"Balance sheet is OUT OF BALANCE by ${abs(balance_difference):,.2f}.",
            _PDF_META,
        ),
        PageBreak(),
    ]

    # ---- Cash Flow
    story += [
        Paragraph("Statement of Cash Flows", heading_2),
        Paragraph(period_label, _PDF_META),
        Paragraph(
            "Current: " + (
                "Ready - cash ties, operating reconciles, and all cash activity "
                "is classified."
                if comparative_cash_flow["current_ready"] else
                "Review required - see the warnings below before relying on "
                "this statement."
            ) + (
                " Prior year: " + (
                    "Ready." if comparative_cash_flow["prior_ready"] else
                    "Review required."
                )
                if comparative_cash_flow["prior_available"] else ""
            ),
            _PDF_META,
        ),
        Spacer(1, 10),
    ]
    cash_flow_sections = CASH_FLOW_STATEMENT_SECTIONS + ((
        "Unclassified Cash Activity",
        "unclassified",
        "Net Unclassified Cash Activity",
    ),)
    for title, key, total_label in cash_flow_sections:
        section = comparative_cash_flow[key]
        if (
            key == "unclassified"
            and not section["lines"]
            and not section["current_entries"]
            and not section["prior_entries"]
        ):
            continue
        story += [
            Paragraph(title, heading_2),
            _pdf_table(
                ["Line", "Current", "Prior Year", "$ Change", "% Change"],
                [[_wrap(item["name"])] + _pdf_comparison_values(item)
                 for item in section["lines"]]
                or [[f"No {title.lower()} recorded", "", "", "", ""]],
                [4.4 * inch, 1.35 * inch, 1.35 * inch, 1.35 * inch, 1.0 * inch],
                money_from=1,
                totals_row=[total_label] + _pdf_comparison_values(
                    section["total"], totals=True
                ),
            ),
            Spacer(1, 8),
        ]
    cash_rollforward_rows = [
        ["NET CHANGE IN CASH"] + _pdf_comparison_values(
            comparative_cash_flow["computed_cash_change"], totals=True
        ),
    ]
    reconciliation = comparative_cash_flow["reconciliation_difference"]
    if reconciliation["current"] or reconciliation["prior"]:
        cash_rollforward_rows.append(
            ["CASH FLOW RECONCILIATION DIFFERENCE"]
            + _pdf_comparison_values(reconciliation, totals=True)
        )
    cash_rollforward_rows.extend([
        ["CASH AT BEGINNING OF PERIOD"] + _pdf_comparison_values(
            comparative_cash_flow["cash_beginning"], totals=True
        ),
        ["CASH AT END OF PERIOD"] + _pdf_comparison_values(
            comparative_cash_flow["cash_ending"], totals=True
        ),
    ])
    cash_rollforward = _pdf_table(
        ["", "Current", "Prior Year", "$ Change", "% Change"],
        cash_rollforward_rows,
        [4.4 * inch, 1.35 * inch, 1.35 * inch, 1.35 * inch, 1.0 * inch],
        money_from=1,
        bold_data_rows=list(range(len(cash_rollforward_rows))),
        ruled_data_rows=[0, len(cash_rollforward_rows) - 1],
    )
    cash_quality_block = [cash_rollforward]
    current_warnings = comparative_cash_flow["current_warnings"]
    prior_warnings = comparative_cash_flow["prior_warnings"]
    if current_warnings or prior_warnings:
        cash_quality_block += [Spacer(1, 8)] + [
            _safe_paragraph(f"Current warning: {warning}", _PDF_META)
            for warning in current_warnings
        ] + [
            _safe_paragraph(f"Prior-year warning: {warning}", _PDF_META)
            for warning in prior_warnings
        ]
    story.append(KeepTogether(cash_quality_block))
    for period_name, entries in [
        ("Current", comparative_cash_flow["unclassified"]["current_entries"]),
        ("Prior-Year", comparative_cash_flow["unclassified"]["prior_entries"]),
    ]:
        if entries:
            story += [
                Spacer(1, 10),
                Paragraph(f"{period_name} Unclassified Cash Activity Details", heading_2),
                _pdf_table(
                    ["Date", "Entry #", "Reason", "Description", "Accounts", "Amount"],
                    [[
                        str(item["entry_date"]), str(item["entry_id"]),
                        _wrap(item["reason"]),
                        _wrap(item["description"] or "No description"),
                        ", ".join(item["account_numbers"]) or "none",
                        _money(item["amount"]),
                    ] for item in entries],
                    [0.8 * inch, 0.65 * inch, 2.55 * inch, 2.75 * inch,
                     1.5 * inch, 1.0 * inch],
                    money_from=5,
                ),
            ]
    for period_name, entries in [
        ("Current", comparative_cash_flow["current_noncash_items"]),
        ("Prior-Year", comparative_cash_flow["prior_noncash_items"]),
    ]:
        if entries:
            story += [
                Spacer(1, 10),
                Paragraph(
                    f"{period_name} Noncash Investing and Financing Activity",
                    heading_2,
                ),
                _pdf_table(
                    ["Date", "Entry #", "Description", "Accounts", "Amount"],
                    [[item["entry_date"], str(item["entry_id"]),
                      _wrap(item["description"] or "No description"),
                      ", ".join(item["accounts"]), _money(item["amount"])]
                     for item in entries],
                    [0.9 * inch, 0.65 * inch, 4.5 * inch, 2.0 * inch, 1.0 * inch],
                    money_from=4,
                ),
            ]
    story.append(PageBreak())

    # ---- Final Trial Balance
    story.append(Paragraph("Final Trial Balance", heading_2))
    money_w = 0.72 * inch
    comparison_by_number = {
        row['account_number']: row for row in comparative_tb['accounts']
    }
    pdf_tb_rows = []
    current_numbers = set()
    for row in tb_rows:
        current_numbers.add(row.account_number)
        comparison_row = comparison_by_number.get(row.account_number, {})
        pdf_tb_rows.append([
            row.account_number, _wrap(row.account_name),
            _money(row.beginning_dr), _money(row.beginning_cr),
            _money(row.period_debits), _money(row.period_credits),
            _money(row.aje_debits), _money(row.aje_credits),
            _money(row.adjusted_dr), _money(row.adjusted_cr),
            _money(comparison_row.get('prior_debit')),
            _money(comparison_row.get('prior_credit')),
        ])
    for comparison_row in comparative_tb['accounts']:
        if comparison_row['account_number'] in current_numbers:
            continue
        if not (comparison_row.get('prior_debit') or
                comparison_row.get('prior_credit')):
            continue
        pdf_tb_rows.append([
            comparison_row['account_number'], _wrap(comparison_row['name']),
            "", "", "", "", "", "", "", "",
            _money(comparison_row.get('prior_debit')),
            _money(comparison_row.get('prior_credit')),
        ])
    story.append(_pdf_table(
        ["Acct #", "Account Name", "Beg Dr", "Beg Cr", "Activity Dr",
         "Activity Cr", "AJE Dr", "AJE Cr", "Final Dr", "Final Cr",
         "PY Final Dr", "PY Final Cr"],
        pdf_tb_rows,
        [0.55 * inch, 2.25 * inch] + [money_w] * 10, money_from=2,
        totals_row=["", "TOTALS",
                    _money(round(sum(r.beginning_dr for r in tb_rows), 2)),
                    _money(round(sum(r.beginning_cr for r in tb_rows), 2)),
                    _money(round(sum(r.period_debits for r in tb_rows), 2)),
                    _money(round(sum(r.period_credits for r in tb_rows), 2)),
                    _money(round(sum(r.aje_debits for r in tb_rows), 2)),
                    _money(round(sum(r.aje_credits for r in tb_rows), 2)),
                    _money(total_dr), _money(total_cr),
                    _money(comparative_tb['prior_total_debits']),
                    _money(comparative_tb['prior_total_credits'])],
    ))
    story.append(PageBreak())

    # ---- Close Map (annual packages only)
    if close_map is not None:
        story += [
            Paragraph("Close Map", heading_2),
            Paragraph(
                ("All required balances were reviewed and current."
                 if close_map["ready"] else
                 f"{close_map['incomplete_count']} required balances still needed attention."),
                _PDF_META,
            ),
            Spacer(1, 10),
            _pdf_table(
                ["Acct #", "Account", "Adjusted", "PY", "$ Change", "% Change",
                 "Lead", "Evidence", "Notes", "Status", "Prepared", "Reviewed"],
                [[row.account_number, _wrap(row.account_name), _money(row.current_balance),
                  _money(row.prior_balance), _money(row.change),
                  _percent(row.change_percent), row.group_code,
                  str(row.evidence_count), str(row.open_note_count), row.status,
                  row.prepared_by, row.reviewed_by]
                 for row in close_map["rows"]],
                [0.55 * inch, 1.75 * inch, 0.8 * inch, 0.8 * inch, 0.8 * inch,
                 0.65 * inch, 0.45 * inch, 0.55 * inch, 0.45 * inch,
                 0.8 * inch, 1.0 * inch, 1.0 * inch],
                money_from=2,
            ),
            Spacer(1, 14),
            Paragraph("Support and explanations", heading_2),
            _pdf_table(
                ["Account", "Evidence References", "Explanation", "Open Review Notes"],
                [[_wrap(f"{row.account_number} - {row.account_name}"),
                  _wrap(row.evidence_references or "None recorded"),
                  _wrap(row.explanation or "None recorded"),
                  _wrap(row.open_notes or "None")]
                 for row in close_map["rows"]],
                [2.0 * inch, 2.4 * inch, 4.0 * inch, 1.2 * inch],
                money_from=None,
            ),
            PageBreak(),
        ]

    # ---- Transactions
    story.append(Paragraph("Transactions", heading_2))
    if transactions:
        story.append(_pdf_table(
            ["Date", "Entry #", "Type", "Description", "Acct #", "Account",
             "Debit", "Credit", "Memo"],
            [[t["entry_date"], str(t["entry_id"]), _wrap(t["entry_type"]),
              _wrap(t["description"]), t["account_number"],
              _wrap(t["account_name"]), _money(t["debit"]),
              _money(t["credit"]), _wrap(t["memo"])]
             for t in transactions],
            [0.75 * inch, 0.55 * inch, 0.75 * inch, 2.5 * inch, 0.55 * inch,
             1.9 * inch, 0.8 * inch, 0.8 * inch, 1.4 * inch], money_from=6,
        ))
    else:
        story.append(Paragraph("No transactions in this period.", _PDF_BODY))
    story.append(PageBreak())

    # ---- Adjusting Entries
    story.append(Paragraph("Adjusting Journal Entries", heading_2))
    if ajes:
        story.append(_pdf_table(
            ["Date", "Entry #", "AJE Ref", "Description", "Acct #", "Account",
             "Debit", "Credit", "Memo"],
            [[t["entry_date"], str(t["entry_id"]), _wrap(t["source_reference"]),
              _wrap(t["description"]), t["account_number"],
              _wrap(t["account_name"]), _money(t["debit"]),
              _money(t["credit"]), _wrap(t["memo"])]
             for t in ajes],
            [0.75 * inch, 0.55 * inch, 0.75 * inch, 2.5 * inch, 0.55 * inch,
             1.9 * inch, 0.8 * inch, 0.8 * inch, 1.4 * inch], money_from=6,
        ))
    else:
        story.append(Paragraph("No adjusting entries in this period.", _PDF_BODY))

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    buffer.seek(0)
    return buffer
