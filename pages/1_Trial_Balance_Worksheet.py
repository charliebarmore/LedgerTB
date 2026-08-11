"""
Trial Balance Worksheet - Primary CPA working view

This page displays a comprehensive trial balance worksheet with columns for:
- Beginning Balance
- Period Activity (Debits/Credits)
- Unadjusted Trial Balance
- Adjusting Journal Entries (AJEs)
- Adjusted Trial Balance
"""

import streamlit as st
import sys
from pathlib import Path
from datetime import date, datetime
from io import BytesIO

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from database import init_database
from utils.client_selector import render_client_selector, get_selected_client
from utils.ui import apply_default_on_change
from utils.unlock import require_unlock
from utils import icons
from utils.export import set_excel_literal
from models.client import Client
from models.fiscal_period import FiscalPeriod
from models.reports import ReportGenerator
from models.journal_entry import JournalEntry, JournalEntryLine
from models.account import Account
from models.audit_log import AuditLog
from services.close_package import (
    build_close_package,
    build_close_package_pdf,
    consistent_export_window,
    load_close_package_snapshot,
)


st.set_page_config(
    page_title="Trial Balance Worksheet",
    page_icon=icons.TRIAL_BALANCE,
    layout="wide"
)

# Render client selector
# Gate on the database passphrase before any DB access, then ensure schema.
require_unlock()
init_database()

client_id = render_client_selector()

if not client_id:
    st.warning("Please select or create a client first.")
    st.stop()

client = Client.get_by_id(client_id)
if not client:
    st.error("Client not found.")
    st.stop()

st.title("Trial Balance Worksheet")
st.caption(f"Viewing: **{client.name}**")

# Period Selection
st.markdown("---")

col1, col2, col3, col4 = st.columns([2, 1, 1, 1])

with col1:
    # Get current year and ensure periods exist
    current_year = date.today().year
    fiscal_year_end = client.fiscal_year_end_month

    # Check available years from existing periods
    periods = FiscalPeriod.get_all(client_id)

    # Get unique years from period names
    available_years = set()
    for p in periods:
        if p.period_type == "Year":
            try:
                year_str = p.period_name.replace("FY ", "")
                available_years.add(int(year_str))
            except ValueError:
                pass

    # Add current year if not present
    if current_year not in available_years:
        FiscalPeriod.ensure_periods_exist(client_id, current_year, fiscal_year_end)
        available_years.add(current_year)

    # Allow user to add other years (show last 5 years as options).
    # bottom-align so the "Add Year" popover lines up with the selectbox box
    # (it has no label above it, unlike the selectbox).
    year_col1, year_col2 = st.columns([3, 1], vertical_alignment="bottom")

    with year_col1:
        # Sort years descending
        years_list = sorted(available_years, reverse=True)

        selected_year = st.selectbox(
            "Fiscal Year",
            options=years_list,
            index=0,
            key="worksheet_year"
        )

    with year_col2:
        # Add year button with popover for year selection
        with st.popover("+ Add Year"):
            # Show years that aren't already available
            potential_years = [y for y in range(current_year, current_year - 10, -1) if y not in available_years]
            if potential_years:
                new_year = st.selectbox("Select year to add", options=potential_years, key="add_year_select")
                if st.button("Add Fiscal Year", key="add_year_btn"):
                    FiscalPeriod.generate_periods(client_id, new_year, fiscal_year_end)
                    st.success(f"Added FY {new_year}")
                    st.rerun()
            else:
                st.info("All recent years are already available")

    # Ensure periods exist for selected year
    FiscalPeriod.ensure_periods_exist(client_id, selected_year, fiscal_year_end)

# Get periods for selected year
all_periods = FiscalPeriod.get_all(client_id)
year_periods = [p for p in all_periods if f"FY {selected_year}" in p.period_name]

# Sort by period type then by start date
type_order = {"Year": 0, "Quarter": 1, "Month": 2, "Custom": 3}
year_periods.sort(key=lambda x: (type_order.get(x.period_type, 4), x.start_date))

with col2:
    period_options = {p.id: f"{p.period_name} ({p.start_date.strftime('%m/%d/%y')} - {p.end_date.strftime('%m/%d/%y')})" for p in year_periods}

    if not period_options:
        st.warning("No periods available")
        st.stop()

    # Default to year period
    default_period = next((p for p in year_periods if p.period_type == "Year"), year_periods[0])

    if 'selected_period_id' not in st.session_state or st.session_state.get('last_year') != selected_year:
        st.session_state.selected_period_id = default_period.id
        st.session_state.last_year = selected_year

    selected_period_id = st.selectbox(
        "Period",
        options=list(period_options.keys()),
        format_func=lambda x: period_options[x],
        index=list(period_options.keys()).index(st.session_state.selected_period_id) if st.session_state.selected_period_id in period_options else 0,
        key="period_selector"
    )
    st.session_state.selected_period_id = selected_period_id

selected_period = FiscalPeriod.get_by_id(selected_period_id)

# Keyed date inputs ignore value= once they hold state, so picking a new
# Period would silently leave the old dates (and numbers) in place. Re-apply
# the period's dates only when the period actually changes — a hand-edited
# range survives everything else.
apply_default_on_change("period_start", depends_on=selected_period_id,
                        default_value=selected_period.start_date)
apply_default_on_change("period_end", depends_on=selected_period_id,
                        default_value=selected_period.end_date)

with col3:
    period_start = st.date_input(
        "From",
        value=selected_period.start_date,
        key="period_start"
    )

with col4:
    period_end = st.date_input(
        "To",
        value=selected_period.end_date,
        key="period_end"
    )

if period_start > period_end:
    st.error("Worksheet period start date cannot be after the end date.")
    st.stop()

# Show all accounts toggle
show_all = st.checkbox("Show all accounts (default: only accounts with activity)", value=False, key="show_all_accounts")

# Year close / reopen — locks all journal entries dated within the fiscal year
year_period = next((p for p in year_periods if p.period_type == "Year"), None)
if year_period:
    checklist = FiscalPeriod.get_close_checklist(year_period.id, client_id)
    st.subheader("Year-close checklist")
    check_cols = st.columns(4)
    with check_cols[0]:
        if checklist["trial_balance_balanced"]:
            st.success("Trial balance balanced")
        else:
            st.error("Trial balance out of balance")
        # \$ so markdown doesn't treat the paired $…$ as inline LaTeX.
        st.caption(
            f"Debits \\${checklist['total_debits']:,.2f} · "
            f"Credits \\${checklist['total_credits']:,.2f}"
        )
    with check_cols[1]:
        if checklist["pending_imports"]:
            st.warning(f"{checklist['pending_imports']} pending imports")
        else:
            st.success("No pending imports")
    with check_cols[2]:
        if checklist["uncategorized_items"]:
            st.warning(f"{checklist['uncategorized_items']} uncategorized items")
        else:
            st.success("No uncategorized items")
    with check_cols[3]:
        if checklist["unresolved_duplicates"]:
            st.warning(f"{checklist['unresolved_duplicates']} potential duplicates")
        else:
            st.success("No potential duplicates")

    if checklist["warning_count"]:
        st.warning(
            "Outstanding items should be resolved before close. If they are understood and "
            "intentional, the close requires a separate acknowledgement."
        )
        st.page_link("pages/4_Import_Transactions.py", label="Review imports")

    if year_period.is_closed:
        lock_cols = st.columns([3, 1])
        with lock_cols[0]:
            st.warning(f"FY {selected_year} is closed. Entries in this year are locked.")
        with lock_cols[1]:
            if st.button("Reopen year", key="reopen_year", width="stretch"):
                FiscalPeriod.set_closed(year_period.id, False, client_id)
                st.rerun()
    else:
        # Collapsed by default: closing the year is an occasional action, and
        # expanded it pushed the worksheet itself below the fold.
        with st.expander(f"Close fiscal year {selected_year} (FY is open)"):
            confirmation_phrase = f"CLOSE FY {selected_year}"
            close_confirmation = st.text_input(
                f"Type {confirmation_phrase} to confirm",
                key="close_year_confirmation",
                placeholder=confirmation_phrase,
            )
            warnings_acknowledged = not checklist["warning_count"] or st.checkbox(
                "I reviewed the outstanding items and accept closing with these warnings.",
                key="close_warning_acknowledgement",
            )
            explicitly_confirmed = close_confirmation.strip().upper() == confirmation_phrase
            if st.button(
                "Close fiscal year", key="close_year", type="primary",
                disabled=not explicitly_confirmed or not warnings_acknowledged,
            ):
                try:
                    FiscalPeriod.set_closed(
                        year_period.id, True, client_id,
                        confirmation={
                            "explicit_confirmation": explicitly_confirmed,
                            "warnings_acknowledged": warnings_acknowledged,
                        },
                    )
                    st.rerun()
                except ValueError as exc:
                    st.error(str(exc))

st.markdown("---")

# Generate the trial balance worksheet
rows, aje_details = ReportGenerator.trial_balance_worksheet(
    client_id=client_id,
    period_start=period_start,
    period_end=period_end,
    show_all_accounts=show_all
)

if not rows:
    st.info("No transactions found for the selected period. Try selecting 'Show all accounts' or a different period.")

    # Still show action buttons
    col1, col2, col3 = st.columns([1, 1, 4])
    with col1:
        if st.button("+ Add AJE", type="primary", key="add_aje_empty"):
            st.session_state.show_aje_form = True
            st.session_state.aje_prefill_account = None
else:
    # Calculate totals
    total_beg_dr = sum(r.beginning_dr for r in rows)
    total_beg_cr = sum(r.beginning_cr for r in rows)
    total_period_dr = sum(r.period_debits for r in rows)
    total_period_cr = sum(r.period_credits for r in rows)
    total_unadj_dr = sum(r.unadjusted_dr for r in rows)
    total_unadj_cr = sum(r.unadjusted_cr for r in rows)
    total_aje_dr = sum(r.aje_debits for r in rows)
    total_aje_cr = sum(r.aje_credits for r in rows)
    total_adj_dr = sum(r.adjusted_dr for r in rows)
    total_adj_cr = sum(r.adjusted_cr for r in rows)

    # Build a clean trial-balance table. Numbers are pre-formatted to strings
    # (commas, 2 decimals) with a dash for empty cells — Streamlit's dataframe
    # ignores a Styler's na_rep, so we format here rather than leave blanks.
    import pandas as pd

    def amt(v):
        return f"{v:,.2f}" if v and v > 0 else "-"

    def make_row(acct, name, beg_dr, beg_cr, act_dr, act_cr, un_dr, un_cr, aje_dr, aje_cr, adj_dr, adj_cr):
        return {
            "Acct #": acct,
            "Account Name": name,
            "Beg Dr": amt(beg_dr), "Beg Cr": amt(beg_cr),
            "Activity Dr": amt(act_dr), "Activity Cr": amt(act_cr),
            "Unadj Dr": amt(un_dr), "Unadj Cr": amt(un_cr),
            "AJE Dr": amt(aje_dr), "AJE Cr": amt(aje_cr),
            "Adj Dr": amt(adj_dr), "Adj Cr": amt(adj_cr),
        }

    table = [
        make_row(r.account_number, r.account_name,
                 r.beginning_dr, r.beginning_cr, r.period_debits, r.period_credits,
                 r.unadjusted_dr, r.unadjusted_cr, r.aje_debits, r.aje_credits,
                 r.adjusted_dr, r.adjusted_cr)
        for r in rows
    ]
    table.append(make_row(
        "", "TOTALS",
        total_beg_dr, total_beg_cr, total_period_dr, total_period_cr,
        total_unadj_dr, total_unadj_cr, total_aje_dr, total_aje_cr,
        total_adj_dr, total_adj_cr))

    df = pd.DataFrame(table)
    num_cols = ["Beg Dr", "Beg Cr", "Activity Dr", "Activity Cr",
                "Unadj Dr", "Unadj Cr", "AJE Dr", "AJE Cr", "Adj Dr", "Adj Cr"]

    # Right-align the numeric columns (best-effort; Streamlit honors text-align)
    styler = df.style.set_properties(subset=num_cols, **{"text-align": "right"})

    st.dataframe(
        styler,
        hide_index=True,
        width="stretch",
        height=min(len(df) * 36 + 40, 660),
        column_config={"Account Name": st.column_config.TextColumn(width="medium")},
    )

    # Drill into the general ledger for a chosen account (replaces the old
    # clickable account-name buttons)
    acct_label = {r.account_id: f"{r.account_number} - {r.account_name}" for r in rows}
    dd1, dd2 = st.columns([3, 1])
    with dd1:
        gl_pick = st.selectbox(
            "Drill into general ledger",
            options=list(acct_label.keys()),
            format_func=lambda aid: acct_label.get(aid, ""),
            key="gl_drill_select",
        )
    with dd2:
        st.write("")
        if st.button("Open GL →", width="stretch", key="open_gl_btn"):
            st.session_state.gl_account_id = gl_pick
            st.switch_page("pages/5_Reports.py")

    # AJE detail (previously shown inline in the AJE cells)
    if aje_details:
        with st.expander("AJE detail"):
            for aid, ajes in aje_details.items():
                name = acct_label.get(aid, str(aid))
                for aje in ajes:
                    side = (f"${aje['debit']:,.2f} Dr" if aje['debit'] > 0
                            else f"${aje['credit']:,.2f} Cr")
                    st.caption(f"{aje['aje_reference']} · {name}: {aje['description']} — {side}")

    # Balance check indicators
    st.markdown("---")
    check_cols = st.columns(4)

    with check_cols[0]:
        beg_diff = abs(total_beg_dr - total_beg_cr)
        if beg_diff < 0.01:
            st.success("Beginning Balance: Balanced")
        else:
            st.error(f"Beginning Balance: Out of balance by ${beg_diff:,.2f}")

    with check_cols[1]:
        unadj_diff = abs(total_unadj_dr - total_unadj_cr)
        if unadj_diff < 0.01:
            st.success("Unadjusted TB: Balanced")
        else:
            st.error(f"Unadjusted TB: Out of balance by ${unadj_diff:,.2f}")

    with check_cols[2]:
        aje_diff = abs(total_aje_dr - total_aje_cr)
        if aje_diff < 0.01:
            st.success("AJEs: Balanced")
        else:
            st.error(f"AJEs: Out of balance by ${aje_diff:,.2f}")

    with check_cols[3]:
        adj_diff = abs(total_adj_dr - total_adj_cr)
        if adj_diff < 0.01:
            st.success("Adjusted TB: Balanced")
        else:
            st.error(f"Adjusted TB: Out of balance by ${adj_diff:,.2f}")

st.markdown("---")

# Action buttons
btn_cols = st.columns([1, 1, 1, 1, 3])

with btn_cols[0]:
    if st.button("+ Add AJE", type="primary", key="add_aje_btn"):
        st.session_state.show_aje_form = True
        st.session_state.aje_prefill_account = None

with btn_cols[1]:
    # Export to Excel with formulas
    if rows:
        # Create Excel with formulas
        output = BytesIO()
        with st.spinner("Generating Excel..."):
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Trial Balance Worksheet"

            # Header
            set_excel_literal(ws['A1'], f"Trial Balance Worksheet - {client.name}")
            ws['A1'].font = Font(bold=True, size=14)
            set_excel_literal(
                ws['A2'],
                f"Period: {period_start.strftime('%m/%d/%Y')} - {period_end.strftime('%m/%d/%Y')}",
            )
            set_excel_literal(
                ws['A3'], f"Generated: {datetime.now().strftime('%m/%d/%Y %H:%M')}",
            )

            # Column headers starting at row 5
            headers = ['Acct #', 'Account Name', 'Type', 'Beg Bal Dr', 'Beg Bal Cr',
                       'Debits', 'Credits', 'Unadj TB Dr', 'Unadj TB Cr',
                       'AJE Dr', 'AJE Cr', 'Adj TB Dr', 'Adj TB Cr']

            for col_idx, header in enumerate(headers, 1):
                cell = set_excel_literal(ws.cell(row=5, column=col_idx), header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            # Data rows starting at row 6
            data_start_row = 6
            for row_idx, row in enumerate(rows, data_start_row):
                set_excel_literal(ws.cell(row=row_idx, column=1), row.account_number)
                set_excel_literal(ws.cell(row=row_idx, column=2), row.account_name)
                set_excel_literal(ws.cell(row=row_idx, column=3), row.account_type)
                ws.cell(row=row_idx, column=4, value=row.beginning_dr if row.beginning_dr > 0 else None)
                ws.cell(row=row_idx, column=5, value=row.beginning_cr if row.beginning_cr > 0 else None)
                ws.cell(row=row_idx, column=6, value=row.period_debits if row.period_debits > 0 else None)
                ws.cell(row=row_idx, column=7, value=row.period_credits if row.period_credits > 0 else None)
                # Unadjusted TB uses formulas
                ws.cell(row=row_idx, column=8, value=f"=MAX(D{row_idx}-E{row_idx}+F{row_idx}-G{row_idx},0)")
                ws.cell(row=row_idx, column=9, value=f"=MAX(E{row_idx}-D{row_idx}+G{row_idx}-F{row_idx},0)")
                ws.cell(row=row_idx, column=10, value=row.aje_debits if row.aje_debits > 0 else None)
                ws.cell(row=row_idx, column=11, value=row.aje_credits if row.aje_credits > 0 else None)
                # Adjusted TB uses formulas
                ws.cell(row=row_idx, column=12, value=f"=MAX(H{row_idx}-I{row_idx}+J{row_idx}-K{row_idx},0)")
                ws.cell(row=row_idx, column=13, value=f"=MAX(I{row_idx}-H{row_idx}+K{row_idx}-J{row_idx},0)")

            # Totals row with formulas
            totals_row = data_start_row + len(rows)
            set_excel_literal(ws.cell(row=totals_row, column=1), "TOTALS")
            ws.cell(row=totals_row, column=1).font = Font(bold=True)

            for col_idx in range(4, 14):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                ws.cell(row=totals_row, column=col_idx, value=f"=SUM({col_letter}{data_start_row}:{col_letter}{totals_row-1})")
                ws.cell(row=totals_row, column=col_idx).font = Font(bold=True)

            # Format number columns
            for row in ws.iter_rows(min_row=data_start_row, max_row=totals_row, min_col=4, max_col=13):
                for cell in row:
                    cell.number_format = '#,##0.00'

            # Adjust column widths
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 10
            for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
                ws.column_dimensions[col].width = 12

            wb.save(output)
            output.seek(0)

        st.download_button(
            label="Export Excel",
            data=output,
            file_name=f"TB_Worksheet_{client.name}_{period_end.strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            on_click=AuditLog.log_event,
            args=(client_id, "EXPORT", "trial_balance_worksheet_export", {
                "format": "xlsx", "period_start": period_start,
                "period_end": period_end, "row_count": len(rows),
            }),
        )

with btn_cols[2]:
    # Everything needed to hand off a finished period: financial statements,
    # final TB, all transactions, AJEs, and cash-account activity.
    # PDF is the file/record copy; the Excel workbook is for further work.
    # (Replaced the old attest-claw bridge export.)
    if rows:
        with consistent_export_window():
            export_rows, _ = ReportGenerator.trial_balance_worksheet(
                client_id, period_start, period_end
            )
            export_snapshot = load_close_package_snapshot(
                client_id, period_start, period_end
            )
            pdf = build_close_package_pdf(
                client_id, client.name, period_start, period_end, export_rows,
                snapshot=export_snapshot,
            )
            package = build_close_package(
                client_id, client.name, period_start, period_end, export_rows,
                snapshot=export_snapshot,
            )
        st.download_button(
            label="Close Package (PDF)",
            data=pdf,
            file_name=f"ClosePackage_{client.name}_{period_end.strftime('%Y%m%d')}.pdf",
            mime="application/pdf",
            help="One PDF: summary with tie-outs, income statement, balance "
                 "sheet, final trial balance, transactions, adjusting entries, "
                 "receipts & disbursements",
            on_click=AuditLog.log_event,
            args=(client_id, "EXPORT", "close_package_export", {
                "format": "pdf",
                "period_start": period_start, "period_end": period_end,
                "row_count": len(export_rows),
            }),
        )
        st.download_button(
            label="Close Package (Excel)",
            data=package,
            file_name=f"ClosePackage_{client.name}_{period_end.strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="Same package as an Excel workbook, one sheet per report",
            on_click=AuditLog.log_event,
            args=(client_id, "EXPORT", "close_package_export", {
                "format": "xlsx",
                "period_start": period_start, "period_end": period_end,
                "row_count": len(export_rows),
            }),
        )

with btn_cols[3]:
    if st.button("Refresh", key="refresh_btn"):
        st.rerun()

# AJE Entry Form (modal-like experience)
if st.session_state.get('show_aje_form', False):
    st.markdown("---")
    st.subheader("Add Adjusting Journal Entry")

    # Get next AJE reference
    next_aje_ref = JournalEntry.get_next_aje_reference(client_id, period_start, period_end)

    # Get accounts for selection
    accounts = Account.get_all(client_id)
    account_options = {a.id: f"{a.account_number} - {a.name}" for a in accounts}

    with st.form("aje_form"):
        form_cols = st.columns([1, 2, 1])

        with form_cols[0]:
            aje_ref = st.text_input("AJE Reference", value=next_aje_ref, disabled=True)
            aje_date = st.date_input("Date", value=period_end)

        with form_cols[1]:
            aje_desc = st.text_input("Description", placeholder="Describe the adjusting entry...")

        with form_cols[2]:
            aje_source = st.text_input("Source Reference", placeholder="W/P Reference...")

        st.markdown("**Entry Lines**")

        # Pre-fill first account if specified
        prefill_account = st.session_state.get('aje_prefill_account')

        line_cols = st.columns([3, 2, 2, 2])
        with line_cols[0]:
            st.markdown("**Account**")
        with line_cols[1]:
            st.markdown("**Debit**")
        with line_cols[2]:
            st.markdown("**Credit**")
        with line_cols[3]:
            st.markdown("**Memo**")

        # Line 1
        line1_cols = st.columns([3, 2, 2, 2])
        with line1_cols[0]:
            default_idx = list(account_options.keys()).index(prefill_account) if prefill_account and prefill_account in account_options else 0
            line1_account = st.selectbox("Account 1", options=list(account_options.keys()),
                                          format_func=lambda x: account_options[x],
                                          index=default_idx, key="line1_acct", label_visibility="collapsed")
        with line1_cols[1]:
            line1_debit = st.number_input("Debit 1", min_value=0.0, step=0.01, key="line1_dr", label_visibility="collapsed")
        with line1_cols[2]:
            line1_credit = st.number_input("Credit 1", min_value=0.0, step=0.01, key="line1_cr", label_visibility="collapsed")
        with line1_cols[3]:
            line1_memo = st.text_input("Memo 1", key="line1_memo", label_visibility="collapsed")

        # Line 2
        line2_cols = st.columns([3, 2, 2, 2])
        with line2_cols[0]:
            line2_account = st.selectbox("Account 2", options=list(account_options.keys()),
                                          format_func=lambda x: account_options[x],
                                          index=0, key="line2_acct", label_visibility="collapsed")
        with line2_cols[1]:
            line2_debit = st.number_input("Debit 2", min_value=0.0, step=0.01, key="line2_dr", label_visibility="collapsed")
        with line2_cols[2]:
            line2_credit = st.number_input("Credit 2", min_value=0.0, step=0.01, key="line2_cr", label_visibility="collapsed")
        with line2_cols[3]:
            line2_memo = st.text_input("Memo 2", key="line2_memo", label_visibility="collapsed")

        # Line 3 (optional)
        line3_cols = st.columns([3, 2, 2, 2])
        with line3_cols[0]:
            line3_account = st.selectbox("Account 3", options=[None] + list(account_options.keys()),
                                          format_func=lambda x: account_options[x] if x else "(Optional)",
                                          index=0, key="line3_acct", label_visibility="collapsed")
        with line3_cols[1]:
            line3_debit = st.number_input("Debit 3", min_value=0.0, step=0.01, key="line3_dr", label_visibility="collapsed")
        with line3_cols[2]:
            line3_credit = st.number_input("Credit 3", min_value=0.0, step=0.01, key="line3_cr", label_visibility="collapsed")
        with line3_cols[3]:
            line3_memo = st.text_input("Memo 3", key="line3_memo", label_visibility="collapsed")

        # Line 4 (optional)
        line4_cols = st.columns([3, 2, 2, 2])
        with line4_cols[0]:
            line4_account = st.selectbox("Account 4", options=[None] + list(account_options.keys()),
                                          format_func=lambda x: account_options[x] if x else "(Optional)",
                                          index=0, key="line4_acct", label_visibility="collapsed")
        with line4_cols[1]:
            line4_debit = st.number_input("Debit 4", min_value=0.0, step=0.01, key="line4_dr", label_visibility="collapsed")
        with line4_cols[2]:
            line4_credit = st.number_input("Credit 4", min_value=0.0, step=0.01, key="line4_cr", label_visibility="collapsed")
        with line4_cols[3]:
            line4_memo = st.text_input("Memo 4", key="line4_memo", label_visibility="collapsed")

        submit_cols = st.columns([1, 1, 4])

        with submit_cols[0]:
            submitted = st.form_submit_button("Save AJE", type="primary")

        with submit_cols[1]:
            if st.form_submit_button("Cancel"):
                st.session_state.show_aje_form = False
                st.session_state.aje_prefill_account = None
                st.rerun()

        if submitted:
            # Build journal entry
            lines = []

            if line1_account and (line1_debit > 0 or line1_credit > 0):
                lines.append(JournalEntryLine(
                    account_id=line1_account,
                    debit=line1_debit,
                    credit=line1_credit,
                    memo=line1_memo if line1_memo else None
                ))

            if line2_account and (line2_debit > 0 or line2_credit > 0):
                lines.append(JournalEntryLine(
                    account_id=line2_account,
                    debit=line2_debit,
                    credit=line2_credit,
                    memo=line2_memo if line2_memo else None
                ))

            if line3_account and (line3_debit > 0 or line3_credit > 0):
                lines.append(JournalEntryLine(
                    account_id=line3_account,
                    debit=line3_debit,
                    credit=line3_credit,
                    memo=line3_memo if line3_memo else None
                ))

            if line4_account and (line4_debit > 0 or line4_credit > 0):
                lines.append(JournalEntryLine(
                    account_id=line4_account,
                    debit=line4_debit,
                    credit=line4_credit,
                    memo=line4_memo if line4_memo else None
                ))

            entry = JournalEntry(
                client_id=client_id,
                entry_date=aje_date,
                description=aje_desc,
                source_reference=aje_source if aje_source else None,
                entry_type='Adjusting',
                aje_reference=next_aje_ref,
                lines=lines
            )

            try:
                entry.save()
                st.success(f"AJE {next_aje_ref} saved successfully!")
                st.session_state.show_aje_form = False
                st.session_state.aje_prefill_account = None
                st.rerun()
            except ValueError as e:
                st.error(str(e))
