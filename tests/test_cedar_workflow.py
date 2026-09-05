"""Import -> recurring approval -> reconciliation -> close -> reopen/restore."""

from datetime import date
from io import BytesIO
import re

import openpyxl
import pypdfium2 as pdfium
import pytest

from database import connection as dbc
from models import close_map
from models.audit_log import AuditLog
from models.draft_entry import DraftEntry
from models.fiscal_period import FiscalPeriod
from models.journal_entry import JournalEntry
from models.reconciliation import BankReconciliation
from models.reports import ReportGenerator
from models.transaction import ImportedTransaction
from money import to_cents
from services import mcp_tools
from services.backups import create_backup, restore_backup
from services.close_package import build_close_package, build_close_package_pdf, load_close_package_snapshot
from services.csv_import import CSVImporter
from services.import_identity import classify_import_duplicates, hash_source
from services.posting import post_transaction
from services.recurring_entries import generate_occurrence
from tests.conftest import post_entry
from tests.helpers.cedar import BANK_CSV, JANUARY, JANUARY_BALANCES, create_cedar


def _bank_rows(accounts):
    rows = CSVImporter.parse_csv(BANK_CSV, date_column="Date", description_column="Description",
                                 amount_column="Amount", source_id=hash_source(BANK_CSV.encode()),
                                 source_filename="cedar-january.csv")
    for row in rows:
        row["bank_account_id"] = accounts["cash"]
    return rows


def _assert_january(client_id):
    tb = mcp_tools.trial_balance(client_id, "2026-01-31")
    actual = {r["number"]: (to_cents(r["debit"]), to_cents(r["credit"]))
              for r in tb["accounts"]}
    assert actual == JANUARY_BALANCES
    assert to_cents(tb["total_debits"]) == to_cents(tb["total_credits"]) == 1_370_000
    income = mcp_tools.income_statement(client_id, "2026-01-01", "2026-01-31")
    assert (income["total_revenue"], income["total_expenses"], income["net_income"]) == (2500, 1500, 1000)
    bs = mcp_tools.balance_sheet(client_id, "2026-01-31")
    assert bs["total_assets"] == 12_200
    assert bs["total_liabilities"] == 1200
    assert bs["total_equity"] == 11_000
    assert bs["balanced"]


def test_cedar_import_to_close_and_verified_restore(db, tmp_path):
    client_id, other_id, accounts, schedule = create_cedar()
    post_entry(client_id, JANUARY[0], [(accounts["cash"], 10000, 0),
                                     (accounts["capital"], 0, 10000)],
               entry_type="Beginning Balance")
    rows = _bank_rows(accounts)
    assert [r["source_row_number"] for r in rows] == [2, 3]
    assert classify_import_duplicates(rows, client_id) == 0
    imported_ids = []
    for row, target in zip(rows, [accounts["revenue"], accounts["office"]]):
        entry, imported = post_transaction(client_id, row, target, accounts["cash"], batch_id="cedar-january")
        imported_ids.append((entry.id, imported.id))
    repeated = _bank_rows(accounts)
    assert classify_import_duplicates(repeated, client_id) == 2
    for row, target, expected in zip(repeated, [accounts["revenue"], accounts["office"]], imported_ids):
        entry, imported = post_transaction(client_id, row, target, accounts["cash"], batch_id="retry")
        assert (entry.id, imported.id) == expected
    assert JournalEntry.count(client_id) == 3
    assert len(ImportedTransaction.get_by_status(client_id, "Posted")) == 2

    result = generate_occurrence(client_id, schedule.id, *JANUARY)
    assert generate_occurrence(client_id, schedule.id, *JANUARY)["result"] == "already_generated"
    assert JournalEntry.count(client_id) == 3
    assert mcp_tools.income_statement(client_id, "2026-01-01", "2026-01-31")["net_income"] == 2200
    primary = DraftEntry.get_by_id(result["draft_id"], client_id)
    primary_id = primary.approve()
    with pytest.raises(ValueError, match="pending"):
        primary.approve()
    assert JournalEntry.count(client_id) == 4
    assert DraftEntry.pending_count(client_id) == 1
    _assert_january(client_id)
    assert JournalEntry.count(other_id) == DraftEntry.pending_count(other_id) == 0
    assert mcp_tools.trial_balance(other_id)["total_debits"] == 0

    reconciliation = BankReconciliation.create(client_id, accounts["cash"], *JANUARY, 12_200)
    reconciliation.save_selected_lines([line.line_id for line in reconciliation.lines()])
    assert reconciliation.ledger_balance() == reconciliation.cleared_balance() == 12_200
    assert reconciliation.difference() == 0
    reconciliation.complete()

    # Independently check the machine-readable workbook and rendered PDF.
    tb_rows, _ = ReportGenerator.trial_balance_worksheet(client_id, *JANUARY)
    snapshot = load_close_package_snapshot(client_id, *JANUARY)
    workbook = build_close_package(client_id, "Cedar Demo Services", *JANUARY, tb_rows, snapshot=snapshot)
    wb = openpyxl.load_workbook(BytesIO(workbook.getvalue()))
    income_cells = {row[0]: row[1] for row in wb["Income Statement"].iter_rows(values_only=True)}
    assert income_cells["NET INCOME"] == 1000
    assert income_cells["Total Expenses"] == 1500
    summary = {row[0]: row[1] for row in wb["Summary"].iter_rows(values_only=True)}
    assert summary["Final trial balance - total debits"] == 13_700
    assert wb["Transactions"].max_row == 9  # header + four two-line entries
    pdf = build_close_package_pdf(client_id, "Cedar Demo Services", *JANUARY, tb_rows, snapshot=snapshot)
    with pdfium.PdfDocument(pdf.getvalue()) as document:
        text = "\n".join(page.get_textpage().get_text_range() for page in document)
    assert "Cedar Demo Services" in text
    assert re.search(r"NET INCOME\s+1,000\.00", text)
    assert re.search(r"Total Assets\s+12,200\.00", text)

    reversal = DraftEntry.get_pending(client_id)[0]
    assert reversal.entry_date == "2026-02-01"
    reversal_id = reversal.approve()
    assert JournalEntry.count(client_id) == 5
    assert {JournalEntry.get_by_id(i).aje_reference for i in [primary_id, reversal_id]} == {"AJE-001", "AJE-002"}
    _assert_january(client_id)
    assert mcp_tools.balance_sheet(client_id, "2026-02-01")["total_liabilities"] == 0
    assert mcp_tools.income_statement(client_id, "2026-01-01", "2026-02-01")["total_expenses"] == 300

    # Complete annual supporting review, then ensure restore preserves signoffs.
    year = next(p for p in FiscalPeriod.get_all(client_id) if p.period_type == "Year")
    for account_id in accounts.values():
        close_map.save_explanation(client_id, year.id, account_id, "Agrees to Cedar demo workpaper.")
        close_map.add_evidence(client_id, year.id, account_id, "workpaper", "Cedar-1", "Synthetic evidence")
        close_map.signoff(client_id, year.id, account_id, "preparer")
        close_map.signoff(client_id, year.id, account_id, "reviewer")
    assert close_map.readiness(client_id, year.id)["ready"]
    prior_audit = [log.id for log in AuditLog.get_all(client_id)]
    backup = create_backup(tmp_path / "backups")
    key = dbc.get_active_key()
    dbc.clear_active_key()
    dbc.set_active_key(key)
    dbc.init_database()
    _assert_january(client_id)
    post_entry(client_id, date(2026, 3, 1), [(accounts["cash"], 1, 0), (accounts["revenue"], 0, 1)])
    assert not close_map.readiness(client_id, year.id)["ready"]

    def restore_audit(conn):
        AuditLog.write(conn.cursor(), None, "database_restore", 0, "RESTORE",
                       new_values={"restored_from": "Cedar verified backup"})

    restore_backup(backup.database_path, tmp_path / "backups", audit=restore_audit)
    _assert_january(client_id)
    assert JournalEntry.count(client_id) == 5
    assert DraftEntry.pending_count(client_id) == 0
    assert close_map.readiness(client_id, year.id)["ready"]
    restored_audit = AuditLog.get_all(client_id)
    assert restored_audit[0].action == "RESTORE"
    assert [log.id for log in restored_audit[1:]] == prior_audit
    assert BankReconciliation.get_by_id(reconciliation.id, client_id).status == "Completed"
    with dbc.get_cursor() as cur:
        assert cur.execute("SELECT COUNT(*) FROM recurring_occurrence_drafts").fetchone()[0] == 2
        assert cur.execute("SELECT COUNT(*) FROM audit_log WHERE action='RESTORE'").fetchone()[0] == 1
        assert cur.execute("PRAGMA integrity_check").fetchone()[0] == "ok"
