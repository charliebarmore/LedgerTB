"""Firm/client branding must persist, validate, and land on deliverables."""
from datetime import date
from io import BytesIO

import openpyxl
import pypdfium2 as pdfium
import pytest
import streamlit as st
from streamlit.testing.v1 import AppTest

from models.reports import ReportGenerator
from services.branding import (
    get_branding,
    get_client_branding,
    normalize_hex,
    pending_client_branding_count,
    propose_client_branding,
    resolve_client_branding_proposal,
    save_branding,
    save_client_branding,
)
from services.close_package import build_close_package, build_close_package_pdf
from tests.conftest import page_path, post_entry

def _make_png() -> bytes:
    from PIL import Image as PILImage

    buffer = BytesIO()
    PILImage.new("RGB", (40, 12), "#B85439").save(buffer, "PNG")
    return buffer.getvalue()


_PNG = _make_png()


def test_normalize_hex():
    assert normalize_hex("#b85439") == "#B85439"
    assert normalize_hex("B85439") == "#B85439"
    assert normalize_hex("red") == ""
    assert normalize_hex("") == ""


def test_branding_round_trip_with_logo(db):
    assert get_branding().is_branded is False
    save_branding("Meridian Ledger CPA PC", tagline="Riverton, GA",
                  accent_hex="#B85439", logo=_PNG, logo_mime="image/png")
    branding = get_branding()
    assert branding.firm_name == "Meridian Ledger CPA PC"
    assert branding.accent_hex == "#B85439"
    assert branding.logo == _PNG

    # Re-saving without a logo keeps the stored one…
    save_branding("Meridian Ledger CPA PC", accent_hex="#B85439")
    assert get_branding().logo == _PNG
    # …and explicit removal drops it.
    save_branding("Meridian Ledger CPA PC", keep_existing_logo=False)
    assert get_branding().logo is None


def test_logo_validation(db):
    with pytest.raises(ValueError, match="PNG or JPEG"):
        save_branding("Firm", logo=_PNG, logo_mime="image/svg+xml")
    with pytest.raises(ValueError, match="2MB"):
        save_branding("Firm", logo=b"x" * (2 * 1024 * 1024 + 1),
                      logo_mime="image/png")


def test_client_branding_round_trip_is_client_scoped(client_id):
    assert get_client_branding(client_id).is_branded is False
    save_client_branding(
        client_id, "Northline Studio", "Atlanta, GA · northline.example",
        "#1D434E", _PNG, "image/png",
    )
    branding = get_client_branding(client_id)
    assert branding.display_name == "Northline Studio"
    assert branding.tagline == "Atlanta, GA · northline.example"
    assert branding.accent_hex == "#1D434E"
    assert branding.logo == _PNG

    save_client_branding(client_id, "Northline Studio", accent_hex="#E8913A")
    assert get_client_branding(client_id).logo == _PNG
    save_client_branding(
        client_id, "Northline Studio", keep_existing_logo=False
    )
    assert get_client_branding(client_id).logo is None


def test_assistant_branding_proposal_needs_human_approval(
    client_id, monkeypatch
):
    save_client_branding(
        client_id, "Original Name", "Original line", "#14141A",
        _PNG, "image/png",
    )
    from utils import actor
    monkeypatch.setattr(actor, "_ASSISTANT", True)
    proposal_id = propose_client_branding(
        client_id, display_name="Northline Studio", accent_hex="#1D434E",
        rationale="Matched the client's supplied brand guide.",
    )
    assert pending_client_branding_count(client_id) == 1
    assert get_client_branding(client_id).display_name == "Original Name"
    with pytest.raises(PermissionError, match="cannot approve"):
        resolve_client_branding_proposal(client_id, proposal_id, True)

    monkeypatch.setattr(actor, "_ASSISTANT", False)
    accepted = resolve_client_branding_proposal(client_id, proposal_id, True)
    assert accepted.display_name == "Northline Studio"
    assert accepted.tagline == "Original line"
    assert accepted.accent_hex == "#1D434E"
    assert accepted.logo == _PNG
    assert pending_client_branding_count(client_id) == 0


def test_firm_settings_renders_selected_client_branding(
    client_id, monkeypatch
):
    import utils.client_selector as selector

    monkeypatch.setattr(selector, "render_client_selector", lambda: client_id)
    monkeypatch.setattr(st, "page_link", lambda *args, **kwargs: None)
    page = AppTest.from_file(
        page_path("pages/12_Firm_Settings.py"), default_timeout=30
    ).run()

    assert not page.exception
    assert any(
        item.value == "Client deliverable branding" for item in page.subheader
    )
    assert any(
        item.label == "Client display name" for item in page.text_input
    )
    assert any(
        item.label == "Save client branding" for item in page.button
    )


def test_close_package_carries_the_brand(client_id, accounts):
    post_entry(client_id, date(2026, 1, 15),
               [(accounts["cash"], 250, 0), (accounts["revenue"], 0, 250)])
    save_branding("Meridian Ledger CPA PC", tagline="meridianledgercpa.example",
                  accent_hex="#B85439", logo=_PNG, logo_mime="image/png")
    save_client_branding(
        client_id, "Northline Studio", "Atlanta, GA · northline.example",
        "#1D434E", _PNG, "image/png",
    )

    period = (date(2026, 1, 1), date(2026, 3, 31))
    tb_rows, _ = ReportGenerator.trial_balance_worksheet(client_id, *period)

    pdf = build_close_package_pdf(client_id, "Test Co", *period, tb_rows)
    doc = pdfium.PdfDocument(pdf.read())
    try:
        pages = []
        image_count = 0
        for page_index in range(len(doc)):
            page = doc[page_index]
            try:
                text_page = page.get_textpage()
                try:
                    pages.append(text_page.get_text_range())
                finally:
                    text_page.close()
                image_count += len(list(page.get_objects(
                    filter=[pdfium.raw.FPDF_PAGEOBJ_IMAGE]
                )))
            finally:
                page.close()
        text = "\n".join(pages)
        assert "Northline Studio" in text
        assert "Atlanta, GA" in text
        assert "Meridian Ledger CPA PC" in text
        assert "meridianledgercpa.example" in text
        assert image_count >= 2  # client and preparer logos both made the cover
    finally:
        doc.close()

    wb = openpyxl.load_workbook(BytesIO(build_close_package(
        client_id, "Test Co", *period, tb_rows).read()))
    summary = {wb["Summary"].cell(row=i, column=1).value:
               wb["Summary"].cell(row=i, column=2).value
               for i in range(1, wb["Summary"].max_row + 1)}
    assert wb["Summary"]["A1"].value == "Northline Studio"
    assert summary.get("Client") == "Atlanta, GA · northline.example"
    assert summary.get("Prepared by") == "Meridian Ledger CPA PC"
    assert len(wb["Summary"]._images) == 2
    assert wb["Summary"]["A1"].font.color.rgb.endswith("1D434E")
    assert wb["Income Statement"].oddFooter.left.text == "Northline Studio"


def test_unbranded_package_still_builds(client_id, accounts):
    post_entry(client_id, date(2026, 1, 15),
               [(accounts["cash"], 250, 0), (accounts["revenue"], 0, 250)])
    period = (date(2026, 1, 1), date(2026, 3, 31))
    tb_rows, _ = ReportGenerator.trial_balance_worksheet(client_id, *period)
    pdf = build_close_package_pdf(client_id, "Test Co", *period, tb_rows)
    assert pdf.read().startswith(b"%PDF")
