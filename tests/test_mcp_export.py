"""The file-export seam to workpaper tools (LedgerPDF pairing).

Contract: exports only land inside user-approved roots, work at read level
(with the action audit-logged), and the Excel carries computed totals — not
uncalculated =SUM() formulas that render as literal text in any tool without
a calc engine.
"""
import os
from datetime import date
from io import BytesIO
from pathlib import Path

import openpyxl
import pytest

from database import connection as dbconn
from models.audit_log import AuditLog
from models.client import Client
from services import mcp_tools
from tests.conftest import post_entry


def _seed(client_id, accounts):
    post_entry(client_id, date(2026, 1, 15),
               [(accounts["cash"], 500, 0), (accounts["revenue"], 0, 500)])
    post_entry(client_id, date(2026, 2, 3),
               [(accounts["expense"], 120, 0), (accounts["cash"], 0, 120)])


def test_export_refused_without_roots(client_id, accounts, tmp_path, monkeypatch):
    _seed(client_id, accounts)
    monkeypatch.delenv("LEDGERTB_MCP_EXPORT_ROOTS", raising=False)
    monkeypatch.delenv("PROBOOKS_MCP_EXPORT_ROOTS", raising=False)
    with pytest.raises(ValueError, match="export is off"):
        mcp_tools.export_close_package(client_id, "2026-01-01", "2026-03-31",
                                       str(tmp_path))


def test_export_refused_outside_roots(client_id, accounts, tmp_path, monkeypatch):
    _seed(client_id, accounts)
    approved = tmp_path / "approved"
    approved.mkdir()
    elsewhere = tmp_path / "elsewhere"
    monkeypatch.setenv("LEDGERTB_MCP_EXPORT_ROOTS", str(approved))
    with pytest.raises(ValueError, match="outside"):
        mcp_tools.export_close_package(client_id, "2026-01-01", "2026-03-31",
                                       str(elsewhere))
    with pytest.raises(ValueError, match="outside"):
        mcp_tools.export_close_package(client_id, "2026-01-01", "2026-03-31",
                                       str(approved / ".." / "elsewhere"))


def test_export_writes_both_files_at_read_level(client_id, accounts, tmp_path,
                                                monkeypatch):
    _seed(client_id, accounts)
    monkeypatch.setenv("LEDGERTB_MCP_EXPORT_ROOTS", str(tmp_path))
    monkeypatch.setattr(dbconn, "ASSISTANT_ACCESS_LEVEL", "read")

    result = mcp_tools.export_close_package(
        client_id, "2026-01-01", "2026-03-31", str(tmp_path / "binder-src"))

    pdf = open(result["pdf"], "rb").read()
    assert pdf.startswith(b"%PDF")

    wb = openpyxl.load_workbook(BytesIO(open(result["xlsx"], "rb").read()))
    assert set(wb.sheetnames) >= {"Summary", "Trial Balance", "Transactions"}

    # Gap-2 regression: the TOTALS row must be numbers, never formula text —
    # openpyxl caches no results, so "=SUM(...)" reads as a literal string in
    # any tool without a calc engine.
    tb = wb["Trial Balance"]
    totals_row = tb.max_row
    assert tb.cell(row=totals_row, column=1).value == "TOTALS"
    for col in range(4, 12):
        value = tb.cell(row=totals_row, column=col).value
        assert not (isinstance(value, str) and value.startswith("=")), \
            f"col {col} is an uncalculated formula"
        assert isinstance(value, (int, float))
    assert tb.cell(row=totals_row, column=10).value == pytest.approx(500.0)

    # The export is audit-logged even at read level.
    monkeypatch.setattr(dbconn, "ASSISTANT_ACCESS_LEVEL", None)
    counts = AuditLog.get_filtered_counts(client_id)
    assert counts["total"] > 0


def test_the_folder_chosen_in_the_app_beats_the_environment(client_id, accounts,
                                                            tmp_path,
                                                            monkeypatch):
    """The user's in-app consent must win. An MCP server's environment comes
    from the client's config file, which a shell-capable assistant can edit —
    if the env var overrode the vault, the assistant could rewrite the very
    boundary the user set for it."""
    from utils.secure_store import set_secret
    from utils.assistant_access import credential_names

    _seed(client_id, accounts)
    monkeypatch.delenv("LEDGERTB_MCP_EXPORT_ROOTS", raising=False)
    monkeypatch.delenv("PROBOOKS_MCP_EXPORT_ROOTS", raising=False)

    vault_root = tmp_path / "vault-root"
    vault_root.mkdir()
    names = credential_names(dbconn.DATABASE_PATH)
    set_secret(names.export_roots, str(vault_root))
    result = mcp_tools.export_close_package(
        client_id, "2026-01-01", "2026-03-31", str(vault_root / "julyco"))
    assert result["pdf"].startswith(str(vault_root))

    # An attacker-set environment root does not widen the boundary...
    env_root = tmp_path / "env-root"
    env_root.mkdir()
    monkeypatch.setenv("LEDGERTB_MCP_EXPORT_ROOTS", str(env_root))
    with pytest.raises(ValueError, match="outside"):
        mcp_tools.export_close_package(
            client_id, "2026-01-01", "2026-03-31", str(env_root / "escaped"))
    # ...and the chosen folder keeps working while it is set.
    again = mcp_tools.export_close_package(
        client_id, "2026-01-01", "2026-03-31", str(vault_root / "again"))
    assert again["pdf"].startswith(str(vault_root))


def test_export_refuses_to_write_through_a_planted_symlink(client_id, accounts,
                                                           tmp_path, monkeypatch):
    """The filename is predictable from client and period, so on a shared
    engagement folder a colleague could pre-place a symlink to catch the
    close package. Containment checks the directory; this checks the file."""
    from utils.secure_store import set_secret
    from utils.assistant_access import credential_names

    _seed(client_id, accounts)
    monkeypatch.delenv("LEDGERTB_MCP_EXPORT_ROOTS", raising=False)
    monkeypatch.delenv("PROBOOKS_MCP_EXPORT_ROOTS", raising=False)

    root = tmp_path / "engagement"
    root.mkdir()
    names = credential_names(dbconn.DATABASE_PATH)
    set_secret(names.export_roots, str(root))

    client = Client.get_by_id(client_id)
    stem = f"{client.name} close package 2026-01-01 to 2026-03-31"
    elsewhere = tmp_path / "colleague-copy.pdf"
    (root / f"{stem}.pdf").symlink_to(elsewhere)

    with pytest.raises(ValueError, match="symbolic link"):
        mcp_tools.export_close_package(
            client_id, "2026-01-01", "2026-03-31", str(root))
    assert not elsewhere.exists(), "export was written through the symlink"


@pytest.mark.skipif(os.name == "nt",
                    reason="Windows has no POSIX file mode; NTFS inherits the "
                           "user-profile ACL instead")
def test_exports_are_not_world_readable(client_id, accounts, tmp_path,
                                        monkeypatch):
    """The book is 0600; the unencrypted close package of that same book
    must not be looser."""
    import stat

    from utils.secure_store import set_secret
    from utils.assistant_access import credential_names

    _seed(client_id, accounts)
    monkeypatch.delenv("LEDGERTB_MCP_EXPORT_ROOTS", raising=False)
    monkeypatch.delenv("PROBOOKS_MCP_EXPORT_ROOTS", raising=False)

    root = tmp_path / "exports"
    root.mkdir()
    names = credential_names(dbconn.DATABASE_PATH)
    set_secret(names.export_roots, str(root))

    result = mcp_tools.export_close_package(
        client_id, "2026-01-01", "2026-03-31", str(root / "q1"))
    for path in (result["pdf"], result["xlsx"]):
        mode = stat.S_IMODE(Path(path).stat().st_mode)
        assert mode == 0o600, f"{path} is {oct(mode)}"


def test_probooks_export_root_env_alias_remains_supported(client_id, accounts,
                                                          tmp_path,
                                                          monkeypatch):
    _seed(client_id, accounts)
    monkeypatch.delenv("LEDGERTB_MCP_EXPORT_ROOTS", raising=False)
    monkeypatch.setenv("PROBOOKS_MCP_EXPORT_ROOTS", str(tmp_path))

    result = mcp_tools.export_close_package(
        client_id, "2026-01-01", "2026-03-31", str(tmp_path / "legacy-config"))
    assert result["pdf"].startswith(str(tmp_path))
