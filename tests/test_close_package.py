"""The close package must tie to the books it summarizes."""
from datetime import date
from io import BytesIO

import openpyxl
import pytest

import pypdfium2 as pdfium

from models.reports import ReportGenerator
from models.journal_entry import JournalEntry, JournalEntryLine
from services.close_package import (
    build_close_package,
    build_close_package_pdf,
    get_cash_activity,
    get_period_transactions,
    load_close_package_snapshot,
)
from tests.conftest import post_entry


Q1 = (date(2026, 1, 1), date(2026, 3, 31))


@pytest.fixture
def booked_period(client_id, accounts):
    post_entry(
        client_id, date(2025, 12, 31),
        [(accounts["cash"], 500, 0), (accounts["equity"], 0, 500)],
    )
    post_entry(
        client_id, date(2026, 1, 15),
        [(accounts["cash"], 250, 0), (accounts["revenue"], 0, 250)],
    )
    post_entry(
        client_id, date(2026, 2, 10),
        [(accounts["expense"], 40, 0), (accounts["cash"], 0, 40)],
    )
    post_entry(
        client_id, date(2026, 3, 31),
        [(accounts["expense"], 10, 0), (accounts["revenue"], 10, 0),
         (accounts["cash"], 0, 20)],
        entry_type="Adjusting", source_reference="AJE-1",
    )
    return client_id


def test_period_transactions_are_scoped_and_ordered(booked_period, accounts):
    rows = get_period_transactions(booked_period, *Q1)
    # 2026-12-31 entry excluded; three in-period entries = 2 + 2 + 3 lines
    assert len(rows) == 7
    assert rows[0]["entry_date"] == "2026-01-15"
    assert all("2025" not in r["entry_date"] for r in rows)


def test_cash_activity_walks_beginning_to_ending(booked_period, accounts):
    cash_rows = get_cash_activity(booked_period, *Q1)
    row = next(r for r in cash_rows if r.account_name.lower().startswith("cash"))
    assert row.beginning == 500.0
    assert row.receipts == 250.0
    assert row.disbursements == 60.0
    assert row.ending == 690.0


def test_workbook_sheets_and_tie_outs(booked_period, accounts):
    client_id = booked_period
    tb_rows, _aje_details = ReportGenerator.trial_balance_worksheet(client_id, *Q1)
    package = build_close_package(client_id, "Test Co", *Q1, tb_rows)
    wb = openpyxl.load_workbook(BytesIO(package.read()))

    assert wb.sheetnames == [
        "Summary", "Income Statement", "Balance Sheet", "Trial Balance",
        "Transactions", "Adjusting Entries", "Receipts & Disbursements",
    ]

    # Core financial statements are part of the package, not separate exports.
    income = wb["Income Statement"]
    income_values = {
        income.cell(row=i, column=1).value: income.cell(row=i, column=2).value
        for i in range(1, income.max_row + 1)
    }
    assert income_values["Total Revenue"] == pytest.approx(240.0)
    assert income_values["Total Expenses"] == pytest.approx(50.0)
    assert income_values["NET INCOME"] == pytest.approx(190.0)

    balance = wb["Balance Sheet"]
    balance_values = {
        balance.cell(row=i, column=1).value: balance.cell(row=i, column=2).value
        for i in range(1, balance.max_row + 1)
    }
    assert balance_values["Total Assets"] == pytest.approx(690.0)
    assert balance_values["TOTAL LIABILITIES & EQUITY"] == pytest.approx(690.0)
    assert balance_values["BALANCE CHECK"] == pytest.approx(0.0)

    # Every supporting-table header must fit its Excel column. This guards the
    # close package against unreadable Activity/AJE/receipt headers.
    for sheet_name in [
        "Trial Balance", "Transactions", "Adjusting Entries",
        "Receipts & Disbursements",
    ]:
        sheet = wb[sheet_name]
        for cell in sheet[1]:
            required = min(len(str(cell.value)) + 2, 40)
            actual = sheet.column_dimensions[cell.column_letter].width
            assert actual >= required, (
                f"{sheet_name}!{cell.coordinate} is clipped: {actual} < {required}"
            )

    # Transactions sheet: 7 in-period lines under 1 header row
    tx = wb["Transactions"]
    assert tx.max_row == 8

    # AJE sheet: only the Adjusting entry's 3 lines, tagged with its reference
    aje = wb["Adjusting Entries"]
    assert aje.max_row == 4
    assert all(aje.cell(row=i, column=3).value == "AJE-1" for i in range(2, 5))

    # Trial Balance sheet: one row per account plus header and totals
    tb = wb["Trial Balance"]
    assert tb.max_row == len(tb_rows) + 2
    assert tb.cell(row=tb.max_row, column=1).value == "TOTALS"

    # Summary declares balance
    summary = wb["Summary"]
    cells = {summary.cell(row=i, column=1).value: summary.cell(row=i, column=2).value
             for i in range(1, summary.max_row + 1)}
    assert cells["In balance"] == "YES"
    assert cells["Final trial balance — total debits"] == pytest.approx(
        sum(r.adjusted_dr for r in tb_rows)
    )


def test_close_package_includes_line_by_line_prior_year_comparisons(
    client_id, accounts
):
    post_entry(client_id, date(2025, 1, 15), [
        (accounts["cash"], 100, 0), (accounts["revenue"], 0, 100),
    ])
    post_entry(client_id, date(2025, 2, 15), [
        (accounts["expense"], 20, 0), (accounts["cash"], 0, 20),
    ])
    post_entry(client_id, date(2026, 1, 15), [
        (accounts["cash"], 150, 0), (accounts["revenue"], 0, 150),
    ])
    post_entry(client_id, date(2026, 2, 15), [
        (accounts["expense"], 30, 0), (accounts["cash"], 0, 30),
    ])
    tb_rows, _ = ReportGenerator.trial_balance_worksheet(client_id, *Q1)
    snapshot = load_close_package_snapshot(client_id, *Q1)
    package = build_close_package(
        client_id, "Test Co", *Q1, tb_rows, snapshot=snapshot
    )
    wb = openpyxl.load_workbook(BytesIO(package.read()), data_only=False)

    income = wb["Income Statement"]
    assert [income.cell(4, col).value for col in range(2, 6)] == [
        "2026-01-01 to 2026-03-31", "2025-01-01 to 2025-03-31",
        "$ Change", "% Change",
    ]
    income_rows = {
        income.cell(row, 1).value: [income.cell(row, col).value for col in range(2, 6)]
        for row in range(1, income.max_row + 1)
    }
    assert income_rows["Total Revenue"] == [150, 100, 50, 50]
    assert income_rows["NET INCOME"] == [120, 80, 40, 50]

    balance = wb["Balance Sheet"]
    balance_rows = {
        balance.cell(row, 1).value: [balance.cell(row, col).value for col in range(2, 6)]
        for row in range(1, balance.max_row + 1)
    }
    assert balance_rows["Total Assets"][:3] == [200, 80, 120]

    trial = wb["Trial Balance"]
    assert trial.cell(1, 12).value == "PY Final Dr"
    assert trial.cell(1, 13).value == "PY Final Cr"
    cash_row = next(
        row for row in range(2, trial.max_row)
        if trial.cell(row, 1).value == "1000"
    )
    assert trial.cell(cash_row, 12).value == 80

    pdf = build_close_package_pdf(
        client_id, "Test Co", *Q1, tb_rows, snapshot=snapshot
    )
    doc = pdfium.PdfDocument(pdf.read())
    try:
        text = "\n".join(
            doc[index].get_textpage().get_text_range()
            for index in range(len(doc))
        )
        assert "Prior Year" in text
        assert "Prior period:" in text
        assert "PY Final Dr" in text
    finally:
        doc.close()


def test_annual_close_package_includes_close_map(client_id, accounts):
    from models import close_map
    from models.fiscal_period import FiscalPeriod

    period = FiscalPeriod(
        client_id=client_id, period_name="FY 2026", period_type="Year",
        start_date=date(2026, 1, 1), end_date=date(2026, 12, 31),
    )
    period.save()
    post_entry(
        client_id, date(2026, 2, 1),
        [(accounts["cash"], 250, 0), (accounts["revenue"], 0, 250)],
    )
    close_map.save_explanation(
        client_id, period.id, accounts["cash"], "Agrees to year-end support."
    )
    close_map.add_evidence(
        client_id, period.id, accounts["cash"], "workpaper", "A-1",
        "Year-end bank reconciliation",
    )
    close_map.signoff(client_id, period.id, accounts["cash"], "preparer")
    close_map.signoff(client_id, period.id, accounts["cash"], "reviewer")

    rows, _ = ReportGenerator.trial_balance_worksheet(
        client_id, period.start_date, period.end_date
    )
    workbook = openpyxl.load_workbook(BytesIO(build_close_package(
        client_id, "Test Co", period.start_date, period.end_date, rows
    ).read()))
    assert "Close Map" in workbook.sheetnames
    close_sheet = workbook["Close Map"]
    headers = [cell.value for cell in close_sheet[1]]
    assert "Status" in headers
    cash_row = next(
        row for row in close_sheet.iter_rows(values_only=True) if row[0] == "1000"
    )
    assert "Reviewed" in cash_row
    assert "A-1 (workpaper)" in cash_row


def test_workbook_stores_untrusted_text_as_literals(client_id, accounts):
    dangerous_client = "=CLIENT()"
    dangerous_account = "+CASH()"
    dangerous_description = "-DDE()"
    dangerous_memo = "@SUM(A1:A2)"
    dangerous_source = "=HYPERLINK(\"https://invalid.example\")"

    from models.account import Account
    cash = Account.get_by_id(accounts["cash"], client_id=client_id)
    cash.name = dangerous_account
    cash.save()
    entry = JournalEntry(
        client_id=client_id,
        entry_date=date(2026, 1, 15),
        description=dangerous_description,
        source_reference=dangerous_source,
        entry_type="Adjusting",
        lines=[
            JournalEntryLine(account_id=accounts["cash"], debit=25.0,
                             memo=dangerous_memo),
            JournalEntryLine(account_id=accounts["revenue"], credit=25.0),
        ],
    )
    entry.save()
    tb_rows, _ = ReportGenerator.trial_balance_worksheet(client_id, *Q1)

    wb = openpyxl.load_workbook(BytesIO(
        build_close_package(client_id, dangerous_client, *Q1, tb_rows).read()
    ))
    expected = {
        dangerous_client, dangerous_account, dangerous_description,
        dangerous_memo, dangerous_source,
    }
    found = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value in expected:
                    found[cell.value] = cell.data_type

    assert set(found) == expected
    assert all(data_type == "s" for data_type in found.values())


def test_export_decodes_preescaped_entry_descriptions(client_id, accounts):
    """HTML entities in source descriptions must print as ordinary text.

    The PDF renderer still escapes the normalized value before giving it to
    ReportLab; this only prevents already-escaped source text from being escaped
    twice.  The ledger value itself remains untouched.
    """
    encoded_description = "Defer Lark &amp; Co. Studio advance"
    readable_description = "Defer Lark & Co. Studio advance"
    entry = JournalEntry(
        client_id=client_id,
        entry_date=date(2026, 1, 15),
        description=encoded_description,
        source_reference="AJE-HTML",
        entry_type="Adjusting",
        lines=[
            JournalEntryLine(
                account_id=accounts["cash"], debit=25.0, memo="Lark & Co."
            ),
            JournalEntryLine(account_id=accounts["revenue"], credit=25.0),
        ],
    )
    entry.save()
    tb_rows, _ = ReportGenerator.trial_balance_worksheet(client_id, *Q1)

    workbook = openpyxl.load_workbook(BytesIO(
        build_close_package(client_id, "Test Co", *Q1, tb_rows).read()
    ))
    assert workbook["Transactions"]["D2"].value == readable_description
    assert workbook["Adjusting Entries"]["D2"].value == readable_description
    assert workbook["Transactions"]["I2"].value == "Lark & Co."

    raw_pdf = build_close_package_pdf(
        client_id, "Test Co", *Q1, tb_rows
    ).read()
    document = pdfium.PdfDocument(raw_pdf)
    try:
        text = "\n".join(
            document[index].get_textpage().get_text_range()
            for index in range(len(document))
        )
    finally:
        document.close()
    assert readable_description in text
    assert "&amp;" not in text


def test_pdf_package_contains_every_section(booked_period, accounts):
    client_id = booked_period
    tb_rows, _ = ReportGenerator.trial_balance_worksheet(client_id, *Q1)
    pdf = build_close_package_pdf(client_id, "Test Co", *Q1, tb_rows)
    raw = pdf.read()
    assert raw.startswith(b"%PDF")

    doc = pdfium.PdfDocument(raw)
    try:
        pages = []
        for page_index in range(len(doc)):
            page = doc[page_index]
            try:
                text_page = page.get_textpage()
                try:
                    pages.append(text_page.get_text_range())
                finally:
                    text_page.close()
            finally:
                page.close()
        text = "\n".join(pages)
        for heading in ["Close Package", "Summary", "Cash Activity",
                        "Income Statement", "Balance Sheet",
                        "Final Trial Balance", "Transactions",
                        "Adjusting Journal Entries"]:
            assert heading in text, f"missing section {heading!r}"
        # tie-outs appear in print: cash walk and the AJE reference
        assert "AJE-1" in text
        assert "690.00" in text          # ending cash
        assert "240.00" in text          # total revenue
        assert "190.00" in text          # net income
        assert "Balance sheet is in balance." in text
        assert "TOTALS" in text
        assert len(doc) >= 6
    finally:
        doc.close()


def test_pdf_and_excel_can_share_one_captured_snapshot(
    booked_period, accounts, monkeypatch
):
    import services.close_package as close_package

    tb_rows, _ = ReportGenerator.trial_balance_worksheet(booked_period, *Q1)
    snapshot = load_close_package_snapshot(booked_period, *Q1)

    def unexpected_reload(*args, **kwargs):
        raise AssertionError("the captured snapshot should be reused")

    monkeypatch.setattr(close_package, "get_period_transactions", unexpected_reload)
    monkeypatch.setattr(close_package, "get_cash_activity", unexpected_reload)
    monkeypatch.setattr(close_package, "get_branding", unexpected_reload)
    monkeypatch.setattr(
        close_package.ReportGenerator, "income_statement", unexpected_reload
    )
    monkeypatch.setattr(
        close_package.ReportGenerator, "balance_sheet", unexpected_reload
    )

    xlsx = build_close_package(
        booked_period, "Test Co", *Q1, tb_rows, snapshot=snapshot
    )
    pdf = build_close_package_pdf(
        booked_period, "Test Co", *Q1, tb_rows, snapshot=snapshot
    )
    assert xlsx.read().startswith(b"PK")
    assert pdf.read().startswith(b"%PDF")
